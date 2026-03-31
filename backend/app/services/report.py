"""
数据看板服务
"""
from typing import List
from datetime import date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, and_, distinct

from app.models import User, Course, Exam, Training, ExamRecord, Department, PoliceType, Enrollment, CheckinRecord
from app.models.department import user_departments
from app.models.police_type import user_police_types
from app.schemas.report import (
    KpiResponse, TrendItem, PoliceTypeDistribution, CityRanking,
    TrainingTrendItem, CityAttendanceItem, CityCompletionItem
)
from logger import logger


class ReportService:
    """数据看板服务"""

    def __init__(self, db: Session):
        self.db = db

    def get_kpi(self) -> KpiResponse:
        """获取KPI数据（含培训维度指标）"""
        total_students = self.db.query(User).filter(User.is_active == True).count()
        total_courses = self.db.query(Course).count()
        total_exams = self.db.query(Exam).count()
        total_trainings = self.db.query(Training).count()

        avg_score_result = self.db.query(func.avg(ExamRecord.score)).scalar()
        avg_score = round(float(avg_score_result), 1) if avg_score_result else 0

        total_records = self.db.query(ExamRecord).count()
        pass_records = self.db.query(ExamRecord).filter(ExamRecord.result == 'pass').count()
        pass_rate = round(pass_records / total_records * 100, 1) if total_records > 0 else 0

        # 进行中培训班：按日期动态计算
        today = date.today()
        active_trainings = self.db.query(Training).filter(
            and_(
                Training.start_date <= today,
                Training.end_date >= today
            )
        ).count()

        # 本月参训人数：本月培训班中 approved 学员去重计数
        month_start = today.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        monthly_trainees = self.db.query(func.count(distinct(Enrollment.user_id))).join(
            Training, Enrollment.training_id == Training.id
        ).filter(
            Enrollment.status == 'approved',
            Training.start_date <= month_end,
            Training.end_date >= month_start
        ).scalar() or 0

        # 本月培训完成率：本月结束的培训 / 本月活跃的培训
        monthly_active = self.db.query(Training).filter(
            Training.start_date <= month_end,
            Training.end_date >= month_start
        ).count()
        monthly_completed = self.db.query(Training).filter(
            Training.end_date >= month_start,
            Training.end_date <= month_end
        ).count()
        monthly_completion_rate = round(monthly_completed / monthly_active * 100, 1) if monthly_active > 0 else 0

        # 待审核学员
        pending_enrollments = self.db.query(Enrollment).filter(
            Enrollment.status == 'pending'
        ).count()

        return KpiResponse(
            total_students=total_students,
            total_courses=total_courses,
            total_exams=total_exams,
            total_trainings=total_trainings,
            avg_score=avg_score,
            pass_rate=pass_rate,
            active_trainings=active_trainings,
            monthly_trainees=monthly_trainees,
            monthly_completion_rate=monthly_completion_rate,
            pending_enrollments=pending_enrollments
        )

    def get_trend(self) -> List[TrendItem]:
        """获取月度趋势"""
        results = self.db.query(
            extract('month', ExamRecord.end_time).label('month'),
            func.count(func.distinct(ExamRecord.user_id)).label('students'),
            func.count(ExamRecord.id).label('exams'),
            func.avg(ExamRecord.score).label('avg_score')
        ).filter(
            ExamRecord.end_time.isnot(None)
        ).group_by(
            extract('month', ExamRecord.end_time)
        ).order_by('month').all()

        items = []
        for r in results:
            items.append(TrendItem(
                month=f"{int(r.month)}月",
                students=r.students or 0,
                exams=r.exams or 0,
                avg_score=round(float(r.avg_score), 1) if r.avg_score else 0
            ))
        return items

    def get_training_trend(self) -> List[TrainingTrendItem]:
        """获取近6月培训完成率趋势"""
        today = date.today()
        items = []

        for i in range(5, -1, -1):
            # 计算第 i 个月前的月份
            target = today.replace(day=1) - timedelta(days=i * 30)
            m_start = target.replace(day=1)
            m_end = (m_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            month_label = f"{m_start.month}月"

            # 本月活跃培训总数（时间段与当月有交集）
            total = self.db.query(Training).filter(
                Training.start_date <= m_end,
                Training.end_date >= m_start
            ).count()

            # 本月已结束的培训
            completed = self.db.query(Training).filter(
                Training.end_date >= m_start,
                Training.end_date <= m_end
            ).count()

            rate = round(completed / total * 100, 1) if total > 0 else 0

            items.append(TrainingTrendItem(
                month=month_label,
                completion_rate=rate,
                total=total,
                completed=completed
            ))

        return items

    def get_training_city_attendance(self) -> List[CityAttendanceItem]:
        """各市局本月参训人数"""
        today = date.today()
        month_start = today.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        results = self.db.query(
            Department.name,
            func.count(distinct(Enrollment.user_id)).label('count')
        ).join(
            user_departments, user_departments.c.department_id == Department.id
        ).join(
            Enrollment, Enrollment.user_id == user_departments.c.user_id
        ).join(
            Training, Training.id == Enrollment.training_id
        ).filter(
            Enrollment.status == 'approved',
            Department.is_active == True,
            Training.start_date <= month_end,
            Training.end_date >= month_start
        ).group_by(Department.name).order_by(func.count(distinct(Enrollment.user_id)).desc()).all()

        return [
            CityAttendanceItem(city=r.name, count=r.count)
            for r in results
        ]

    def get_training_city_completion(self) -> List[CityCompletionItem]:
        """各市培训完成率排名"""
        results = self.db.query(
            Department.name,
            func.count(distinct(Enrollment.user_id)).label('total_students')
        ).join(
            user_departments, user_departments.c.department_id == Department.id
        ).join(
            Enrollment, Enrollment.user_id == user_departments.c.user_id
        ).filter(
            Enrollment.status == 'approved',
            Department.is_active == True
        ).group_by(Department.name).all()

        items = []
        for r in results:
            # 该市已结束培训数
            total_trainings = self.db.query(func.count(distinct(Training.id))).join(
                Enrollment, Enrollment.training_id == Training.id
            ).join(
                user_departments, user_departments.c.user_id == Enrollment.user_id
            ).filter(
                user_departments.c.department_id == self.db.query(Department.id).filter(Department.name == r.name).scalar_subquery(),
                Enrollment.status == 'approved'
            ).scalar() or 0

            completed_trainings = self.db.query(func.count(distinct(Training.id))).join(
                Enrollment, Enrollment.training_id == Training.id
            ).join(
                user_departments, user_departments.c.user_id == Enrollment.user_id
            ).filter(
                user_departments.c.department_id == self.db.query(Department.id).filter(Department.name == r.name).scalar_subquery(),
                Enrollment.status == 'approved',
                Training.end_date <= date.today()
            ).scalar() or 0

            rate = round(completed_trainings / total_trainings * 100, 1) if total_trainings > 0 else 0
            items.append(CityCompletionItem(
                city=r.name,
                rate=rate,
                total=total_trainings,
                completed=completed_trainings
            ))

        items.sort(key=lambda x: x.rate, reverse=True)
        return items

    def export_report(self) -> bytes:
        """生成 Excel 报告"""
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.warning("openpyxl not installed, returning empty report")
            return b""

        wb = Workbook()

        # Sheet 1: KPI
        ws = wb.active
        ws.title = "KPI概览"
        kpi = self.get_kpi()
        ws.append(["指标", "数值"])
        ws.append(["进行中培训班", kpi.active_trainings])
        ws.append(["本月参训人数", kpi.monthly_trainees])
        ws.append(["本月培训完成率", f"{kpi.monthly_completion_rate}%"])
        ws.append(["待审核学员", kpi.pending_enrollments])
        ws.append(["注册学员总数", kpi.total_students])
        ws.append(["课程总数", kpi.total_courses])

        # Sheet 2: 培训趋势
        ws2 = wb.create_sheet("培训趋势")
        ws2.append(["月份", "完成率(%)", "活跃培训数", "已完成培训数"])
        for item in self.get_training_trend():
            ws2.append([item.month, item.completion_rate, item.total, item.completed])

        # Sheet 3: 各市参训人数
        ws3 = wb.create_sheet("各市参训人数")
        ws3.append(["地市", "参训人数"])
        for item in self.get_training_city_attendance():
            ws3.append([item.city, item.count])

        # Sheet 4: 各市完成率排名
        ws4 = wb.create_sheet("各市完成率排名")
        ws4.append(["地市", "完成率(%)", "培训总数", "已完成"])
        for item in self.get_training_city_completion():
            ws4.append([item.city, item.rate, item.total, item.completed])

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def get_police_type_distribution(self) -> List[PoliceTypeDistribution]:
        """获取警种分布（通过关联表统计）"""
        results = self.db.query(
            PoliceType.name,
            func.count(user_police_types.c.user_id).label('count')
        ).join(
            user_police_types, PoliceType.id == user_police_types.c.police_type_id
        ).join(
            User, User.id == user_police_types.c.user_id
        ).filter(
            User.is_active == True,
            PoliceType.is_active == True
        ).group_by(PoliceType.name).all()

        return [
            PoliceTypeDistribution(name=r.name, value=r.count)
            for r in results
        ]

    def get_city_ranking(self) -> List[CityRanking]:
        """获取城市/单位排名（通过关联表统计）"""
        results = self.db.query(
            Department.name,
            func.avg(User.avg_score).label('score'),
            func.count(user_departments.c.user_id).label('students')
        ).join(
            user_departments, Department.id == user_departments.c.department_id
        ).join(
            User, User.id == user_departments.c.user_id
        ).filter(
            User.is_active == True,
            Department.is_active == True
        ).group_by(Department.name).order_by(func.avg(User.avg_score).desc()).limit(20).all()

        return [
            CityRanking(
                city=r.name,
                score=round(float(r.score), 1) if r.score else 0,
                students=r.students
            )
            for r in results
        ]
