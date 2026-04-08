"""
考试管理服务
"""
from datetime import datetime
from io import BytesIO
from math import ceil
import secrets
import string
from typing import Any, Dict, List, Optional

from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload, selectinload

from app.models import (
    AdmissionExam,
    AdmissionExamRecord,
    Course,
    Department,
    Enrollment,
    Exam,
    ExamParticipant,
    ExamParticipantImportBatch,
    ExamPaper,
    ExamPaperQuestion,
    ExamRecord,
    KnowledgePoint,
    PaperFolder,
    PoliceType,
    Question,
    Role,
    Training,
    User,
)
from app.schemas import PaginatedResponse
from app.schemas.exam import (
    ADMISSION_SCOPE_ALL,
    ADMISSION_SCOPE_DEPARTMENT,
    ADMISSION_SCOPE_ROLE,
    ADMISSION_SCOPE_USER,
    AdmissionExamCreate,
    AdmissionExamDetailResponse,
    AdmissionExamRecordResponse,
    AdmissionExamResponse,
    AdmissionExamUpdate,
    ExamCreate,
    ExamDetailResponse,
    EXAM_PARTICIPANT_MODE_EXCEL,
    EXAM_PARTICIPANT_MODE_TRAINING,
    EXAM_SCENE_STANDALONE,
    EXAM_SCENE_TRAINING,
    ExamPaperCreate,
    ExamPaperDetailResponse,
    ExamPaperResponse,
    ExamParticipantImportConfirmRequest,
    ExamParticipantImportPreviewResponse,
    ExamParticipantImportRowResponse,
    ExamParticipantResponse,
    ExamPaperUpdate,
    ExamQuestionAnswerDetailResponse,
    ExamQuestionSnapshotResponse,
    ExamRecordResponse,
    ExamResponse,
    ExamSubmit,
    ExamUpdate,
    ExamWrongQuestionResponse,
    PaperFolderCreate,
    PaperFolderResponse,
    PaperFolderUpdate,
)
from app.services.auth import auth_service
from app.utils.authz import (
    can_view_question_with_context,
    can_view_training_with_context,
    is_admin_user,
    is_instructor_user,
)
from app.utils.data_scope import DataScopeContext, build_data_scope_context
from logger import logger


DIMENSION_KEYS = ("law", "enforce", "evidence", "physical", "ethic")
DIMENSION_RULES = {
    "law": ("法律", "法规", "条例", "法"),
    "enforce": ("执法", "程序", "处置", "侦查"),
    "evidence": ("证据", "取证", "笔录"),
    "physical": ("体能", "技能", "警械"),
    "ethic": ("道德", "纪律", "作风", "廉政"),
}
DEFAULT_EXAM_DURATION = 60
DEFAULT_PASSING_SCORE_RATIO = 0.6
EXAM_STATUS_VALUES = {"upcoming", "active", "ended"}
PARTICIPANT_IMPORT_STATUS_MATCHED = "matched"
PARTICIPANT_IMPORT_STATUS_CREATED = "created"
PARTICIPANT_IMPORT_STATUS_FAILED = "failed"


class ExamService:
    """考试服务"""

    def __init__(self, db: Session):
        self.db = db

    def get_exam_papers(
        self,
        page: int = 1,
        size: int = 10,
        status: Optional[str] = None,
        paper_type: Optional[str] = None,
        search: Optional[str] = None,
        current_user_id: Optional[int] = None,
        folder_id: Optional[int] = None,
    ) -> PaginatedResponse[ExamPaperResponse]:
        """获取试卷列表"""
        query = self.db.query(ExamPaper).options(*self._paper_load_options())

        if status:
            query = query.filter(ExamPaper.status == status)
        if paper_type:
            query = query.filter(ExamPaper.type == paper_type)
        if search:
            query = query.filter(ExamPaper.title.contains(search))
        if folder_id is not None:
            query = query.filter(ExamPaper.folder_id == folder_id)

        papers = query.order_by(ExamPaper.created_at.desc(), ExamPaper.id.desc()).all()
        scope_context = self._get_scope_context(current_user_id)
        if scope_context:
            papers = [paper for paper in papers if self._can_view_paper_with_context(scope_context, paper)]
        return self._paginate([self._to_paper_response(paper) for paper in papers], page, size)

    def get_exam_paper_detail(
        self,
        paper_id: int,
        current_user_id: Optional[int] = None,
    ) -> Optional[ExamPaperDetailResponse]:
        """获取试卷详情"""
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            return None
        scope_context = self._get_scope_context(current_user_id)
        if scope_context and not self._can_view_paper_with_context(scope_context, paper):
            return None
        return self._to_paper_detail_response(paper)

    def create_exam_paper(self, data: ExamPaperCreate, user_id: int) -> ExamPaperDetailResponse:
        """创建试卷"""
        questions = self._load_questions(data.question_ids or [])
        resolved_total_score = data.total_score or self._sum_question_scores(questions)
        paper = ExamPaper(
            title=data.title,
            description=data.description,
            duration=data.duration or 60,
            total_score=resolved_total_score,
            passing_score=data.passing_score or 60,
            type=data.type,
            status="draft",
            folder_id=data.folder_id,
            created_by=user_id,
        )
        self.db.add(paper)
        self.db.flush()
        self._replace_paper_questions(paper.id, data.question_ids, questions)
        self.db.commit()

        logger.info("创建试卷: %s", paper.title)
        detail = self.get_exam_paper_detail(paper.id, user_id)
        if not detail:
            raise ValueError("创建试卷后读取详情失败")
        return detail

    def update_exam_paper(self, paper_id: int, data: ExamPaperUpdate) -> ExamPaperDetailResponse:
        """更新试卷"""
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            raise ValueError("试卷不存在")
        if paper.status != "draft":
            raise ValueError("已发布或归档的试卷不能修改")

        update_data = data.model_dump(exclude_unset=True)
        question_ids = update_data.pop("question_ids", None)
        for field, value in update_data.items():
            setattr(paper, field, value)

        if question_ids is not None:
            questions = self._load_questions(question_ids)
            paper.total_score = data.total_score or self._sum_question_scores(questions)
            self._replace_paper_questions(paper.id, question_ids, questions)
        elif data.total_score is not None:
            paper.total_score = data.total_score

        self.db.commit()

        detail = self.get_exam_paper_detail(paper.id, paper.created_by)
        if not detail:
            raise ValueError("试卷不存在")
        return detail

    def publish_exam_paper(self, paper_id: int) -> ExamPaperDetailResponse:
        """发布试卷"""
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            raise ValueError("试卷不存在")
        if not (paper.paper_questions or []):
            raise ValueError("试卷至少需要配置一道题目")
        if paper.status == "published":
            detail = self.get_exam_paper_detail(paper.id, paper.created_by)
            if not detail:
                raise ValueError("试卷不存在")
            return detail

        paper.status = "published"
        paper.published_at = paper.published_at or self._current_time(paper.published_at)
        self.db.commit()

        detail = self.get_exam_paper_detail(paper.id, paper.created_by)
        if not detail:
            raise ValueError("试卷不存在")
        return detail

    def archive_exam_paper(self, paper_id: int) -> ExamPaperDetailResponse:
        """归档试卷"""
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            raise ValueError("试卷不存在")
        if paper.status == "archived":
            detail = self.get_exam_paper_detail(paper.id, paper.created_by)
            if not detail:
                raise ValueError("试卷不存在")
            return detail

        paper.status = "archived"
        self.db.commit()

        detail = self.get_exam_paper_detail(paper.id, paper.created_by)
        if not detail:
            raise ValueError("试卷不存在")
        return detail

    def delete_exam_paper(self, paper_id: int) -> None:
        """删除试卷"""
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            raise ValueError("试卷不存在")
        if self._paper_usage_count(paper) > 0:
            raise ValueError("试卷已被考试引用，不能删除")

        self.db.delete(paper)
        self.db.commit()

    # ============== 文件夹管理 ==============

    def get_paper_folders(self, current_user_id: Optional[int] = None) -> List[PaperFolderResponse]:
        """获取文件夹树"""
        folders = self.db.query(PaperFolder).order_by(PaperFolder.sort_order, PaperFolder.id).all()
        folder_responses = []
        for folder in folders:
            paper_count = self.db.query(ExamPaper).filter(ExamPaper.folder_id == folder.id).count()
            folder_responses.append(PaperFolderResponse(
                id=folder.id,
                name=folder.name,
                parent_id=folder.parent_id,
                sort_order=folder.sort_order,
                paper_count=paper_count,
                children=[],
            ))
        return self._build_folder_tree(folder_responses)

    def _build_folder_tree(self, folders: List[PaperFolderResponse]) -> List[PaperFolderResponse]:
        """构建文件夹树"""
        folder_map = {f.id: f for f in folders}
        root_folders = []
        for folder in folders:
            if folder.parent_id and folder.parent_id in folder_map:
                folder_map[folder.parent_id].children.append(folder)
            else:
                root_folders.append(folder)
        return root_folders

    def create_paper_folder(self, data: PaperFolderCreate, user_id: int) -> PaperFolderResponse:
        """创建文件夹"""
        if data.parent_id:
            parent = self.db.query(PaperFolder).filter(PaperFolder.id == data.parent_id).first()
            if not parent:
                raise ValueError("父文件夹不存在")
        folder = PaperFolder(
            name=data.name,
            parent_id=data.parent_id,
            sort_order=data.sort_order,
            created_by=user_id,
        )
        self.db.add(folder)
        self.db.commit()
        paper_count = self.db.query(ExamPaper).filter(ExamPaper.folder_id == folder.id).count()
        return PaperFolderResponse(
            id=folder.id,
            name=folder.name,
            parent_id=folder.parent_id,
            sort_order=folder.sort_order,
            paper_count=paper_count,
            children=[],
        )

    def update_paper_folder(self, folder_id: int, data: PaperFolderUpdate) -> PaperFolderResponse:
        """更新文件夹"""
        folder = self.db.query(PaperFolder).filter(PaperFolder.id == folder_id).first()
        if not folder:
            raise ValueError("文件夹不存在")
        if data.name is not None:
            folder.name = data.name
        if data.parent_id is not None:
            if data.parent_id == folder_id:
                raise ValueError("不能将文件夹设置为自己的子文件夹")
            parent = self.db.query(PaperFolder).filter(PaperFolder.id == data.parent_id).first()
            if not parent:
                raise ValueError("父文件夹不存在")
            folder.parent_id = data.parent_id
        if data.sort_order is not None:
            folder.sort_order = data.sort_order
        self.db.commit()
        paper_count = self.db.query(ExamPaper).filter(ExamPaper.folder_id == folder.id).count()
        return PaperFolderResponse(
            id=folder.id,
            name=folder.name,
            parent_id=folder.parent_id,
            sort_order=folder.sort_order,
            paper_count=paper_count,
            children=[],
        )

    def delete_paper_folder(self, folder_id: int) -> None:
        """删除文件夹"""
        folder = self.db.query(PaperFolder).filter(PaperFolder.id == folder_id).first()
        if not folder:
            raise ValueError("文件夹不存在")
        # 检查是否有子文件夹
        children = self.db.query(PaperFolder).filter(PaperFolder.parent_id == folder_id).count()
        if children > 0:
            raise ValueError("请先删除子文件夹")
        # 检查是否有试卷
        paper_count = self.db.query(ExamPaper).filter(ExamPaper.folder_id == folder_id).count()
        if paper_count > 0:
            raise ValueError("请先将文件夹内的试卷移出后再删除")
        self.db.delete(folder)
        self.db.commit()

    def move_paper_to_folder(self, paper_id: int, folder_id: Optional[int]) -> ExamPaperDetailResponse:
        """移动试卷到文件夹"""
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            raise ValueError("试卷不存在")
        if folder_id is not None:
            folder = self.db.query(PaperFolder).filter(PaperFolder.id == folder_id).first()
            if not folder:
                raise ValueError("目标文件夹不存在")
        paper.folder_id = folder_id
        self.db.commit()
        detail = self.get_exam_paper_detail(paper.id, paper.created_by)
        if not detail:
            raise ValueError("试卷不存在")
        return detail

    def get_exams(
        self,
        page: int = 1,
        size: int = 10,
        status: Optional[str] = None,
        exam_type: Optional[str] = None,
        search: Optional[str] = None,
        scene: Optional[str] = None,
        training_id: Optional[int] = None,
        purpose: Optional[str] = None,
        department_id: Optional[int] = None,
        police_type_id: Optional[int] = None,
        current_user_id: Optional[int] = None,
    ) -> PaginatedResponse[ExamResponse]:
        """获取统一考试列表"""
        status_filters = self._normalize_exam_status_filters(status)
        current_user = self._get_admission_scope_user(current_user_id)
        query = self.db.query(Exam).options(
            *self._training_exam_load_options(),
        )

        if exam_type:
            query = query.filter(Exam.type == exam_type)
        if scene:
            query = query.filter(Exam.scene == scene)
        if search:
            keyword = str(search).strip()
            if keyword:
                query = query.filter(
                    or_(
                        Exam.title.contains(keyword),
                        Exam.participant_summary.contains(keyword),
                        Exam.training.has(
                            or_(
                                Training.name.contains(keyword),
                                Training.class_code.contains(keyword),
                            )
                        ),
                    )
                )
        if training_id:
            query = query.filter(Exam.training_id == training_id)
        if purpose:
            query = query.filter(Exam.purpose == purpose)

        exams = query.order_by(Exam.created_at.desc(), Exam.id.desc()).all()
        scope_context = self._get_scope_context(current_user_id)
        changed = False
        items: List[ExamResponse] = []
        for exam in exams:
            if self._refresh_exam_status(exam):
                changed = True
            if status_filters and (exam.status or "upcoming") not in status_filters:
                continue
            if department_id and department_id not in self._normalize_scope_target_ids(exam.department_ids):
                continue
            if police_type_id and police_type_id not in self._normalize_scope_target_ids(exam.police_type_ids):
                continue
            if scope_context and (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_TRAINING:
                if not self._can_view_training_exam_with_context(scope_context, exam):
                    continue
            if current_user_id and self._is_student_user(current_user):
                if not self._can_access_exam_as_student(exam, current_user_id):
                    continue
            items.append(self._to_exam_response(exam, current_user_id))

        if changed:
            self.db.commit()

        return self._paginate(items, page, size)

    def get_admission_exams(
        self,
        page: int = 1,
        size: int = 10,
        status: Optional[str] = None,
        exam_type: Optional[str] = None,
        search: Optional[str] = None,
        current_user_id: Optional[int] = None,
    ) -> PaginatedResponse[AdmissionExamResponse]:
        """获取独立准入考试列表"""
        status_filters = self._normalize_exam_status_filters(status)
        current_user = self._get_admission_scope_user(current_user_id)
        can_manage_all = bool(current_user_id and self._can_manage_admission_exam(current_user_id))
        legacy_query = self.db.query(AdmissionExam).options(
            *self._admission_exam_load_options(),
        )
        query = self.db.query(Exam).options(
            *self._training_exam_load_options(),
        ).filter(
            Exam.scene == EXAM_SCENE_STANDALONE,
            Exam.purpose == "admission",
        )

        if exam_type:
            legacy_query = legacy_query.filter(AdmissionExam.type == exam_type)
            query = query.filter(Exam.type == exam_type)
        if search:
            legacy_query = legacy_query.filter(AdmissionExam.title.contains(search))
            query = query.filter(Exam.title.contains(search))

        exams = legacy_query.order_by(AdmissionExam.created_at.desc(), AdmissionExam.id.desc()).all()
        unified_exams = query.order_by(Exam.created_at.desc(), Exam.id.desc()).all()
        changed = False
        items: List[AdmissionExamResponse] = []
        for exam in exams:
            if self._refresh_exam_status(exam):
                changed = True
            if status_filters and (exam.status or "upcoming") not in status_filters:
                continue
            if current_user_id and not can_manage_all:
                if not current_user or not self._can_view_admission_exam(exam, current_user):
                    continue
            items.append(self._to_admission_exam_response(exam, current_user_id, current_user))
        for exam in unified_exams:
            if self._refresh_exam_status(exam):
                changed = True
            if status_filters and (exam.status or "upcoming") not in status_filters:
                continue
            if current_user_id and self._is_student_user(current_user):
                if not self._can_access_exam_as_student(exam, current_user_id):
                    continue
            items.append(self._to_unified_exam_as_admission_response(exam, current_user_id))

        if changed:
            self.db.commit()

        items.sort(key=lambda item: (item.created_at or datetime.min, item.id), reverse=True)
        return self._paginate(items, page, size)

    def _normalize_exam_status_filters(self, status: Optional[str]) -> List[str]:
        if not status:
            return []
        normalized: List[str] = []
        for item in str(status).split(","):
            value = item.strip()
            if not value or value in normalized:
                continue
            if value not in EXAM_STATUS_VALUES:
                continue
            normalized.append(value)
        return normalized

    def create_exam(self, data: ExamCreate, user_id: int) -> ExamResponse:
        """创建统一考试"""
        if data.scene == EXAM_SCENE_TRAINING and data.training_id is None:
            raise ValueError("培训班考试必须关联培训班")
        if data.scene == EXAM_SCENE_STANDALONE:
            data.training_id = None
        if data.scene == EXAM_SCENE_TRAINING and data.training_id is not None:
            training = self.db.query(Training).filter(Training.id == data.training_id).first()
            if not training:
                raise ValueError("关联培训班不存在")
        if data.scene == EXAM_SCENE_STANDALONE and data.participant_mode != EXAM_PARTICIPANT_MODE_EXCEL:
            raise ValueError("独立考试当前仅支持 Excel 名单导入")

        paper = self._get_selectable_paper(data.paper_id, user_id)
        duration = self._resolve_exam_duration(data.duration, paper.duration)
        passing_score = self._resolve_create_exam_passing_score(paper, data.passing_score)
        start_time, end_time = self._validate_exam_configuration(
            total_score=paper.total_score,
            duration=duration,
            passing_score=passing_score,
            start_time=data.start_time,
            end_time=data.end_time,
        )
        exam = Exam(
            paper_id=paper.id,
            title=data.title,
            description=data.description,
            duration=duration,
            total_score=paper.total_score or 0,
            passing_score=passing_score,
            status=data.status,
            type=data.type or paper.type or "formal",
            scene=data.scene,
            participant_mode=data.participant_mode,
            purpose=data.purpose,
            training_id=data.training_id,
            course_ids=self._resolve_selected_course_ids(data.course_ids),
            department_ids=self._normalize_scope_target_ids(data.department_ids),
            police_type_ids=self._normalize_scope_target_ids(data.police_type_ids),
            participant_summary=(data.participant_summary or "").strip() or None,
            max_attempts=data.max_attempts,
            allow_makeup=data.allow_makeup,
            start_time=start_time,
            end_time=end_time,
            published_at=self._current_time(start_time, end_time),
            created_by=user_id,
        )
        self.db.add(exam)
        self.db.commit()
        self.db.refresh(exam)

        logger.info("创建统一考试: %s", exam.title)
        detail = self.get_exam_detail(exam.id, user_id)
        if not detail:
            raise ValueError("创建考试后读取详情失败")
        return detail

    def create_admission_exam(self, data: AdmissionExamCreate, user_id: int) -> AdmissionExamResponse:
        """创建独立准入考试"""
        paper = self._get_selectable_paper(data.paper_id, user_id)
        scope_type, scope_target_ids, scope_summary = self._prepare_admission_scope(
            data.scope_type,
            data.scope_target_ids,
        )
        duration = self._resolve_exam_duration(data.duration, paper.duration)
        passing_score = self._resolve_create_exam_passing_score(paper, data.passing_score)
        start_time, end_time = self._validate_exam_configuration(
            total_score=paper.total_score,
            duration=duration,
            passing_score=passing_score,
            start_time=data.start_time,
            end_time=data.end_time,
        )
        exam = AdmissionExam(
            paper_id=paper.id,
            title=data.title,
            description=data.description,
            duration=duration,
            total_score=paper.total_score or 0,
            passing_score=passing_score,
            status=data.status,
            type=data.type or paper.type or "formal",
            scope=scope_summary,
            scope_type=scope_type,
            scope_target_ids=scope_target_ids,
            course_ids=self._resolve_selected_course_ids(data.course_ids),
            max_attempts=data.max_attempts,
            start_time=start_time,
            end_time=end_time,
            published_at=self._current_time(start_time, end_time),
            created_by=user_id,
        )
        self.db.add(exam)
        self.db.commit()
        self.db.refresh(exam)

        logger.info("创建准入考试: %s", exam.title)
        detail = self.get_admission_exam_detail(exam.id, user_id)
        if not detail:
            raise ValueError("创建准入考试后读取详情失败")
        return detail

    def update_exam(self, exam_id: int, data: ExamUpdate) -> ExamResponse:
        """更新统一考试"""
        exam = self.db.query(Exam).options(
            *self._training_exam_load_options(include_training=False, include_records=False),
        ).filter(Exam.id == exam_id).first()
        if not exam:
            raise ValueError("考试不存在")

        next_scene = data.scene or exam.scene or EXAM_SCENE_TRAINING
        next_training_id = data.training_id if data.training_id is not None else exam.training_id
        next_participant_mode = data.participant_mode or exam.participant_mode or EXAM_PARTICIPANT_MODE_TRAINING
        if next_scene == EXAM_SCENE_TRAINING and not next_training_id:
            raise ValueError("培训班考试必须关联培训班")
        if next_scene == EXAM_SCENE_STANDALONE and next_participant_mode != EXAM_PARTICIPANT_MODE_EXCEL:
            raise ValueError("独立考试当前仅支持 Excel 名单导入")
        if next_scene == EXAM_SCENE_TRAINING and next_training_id is not None:
            training = self.db.query(Training).filter(Training.id == next_training_id).first()
            if not training:
                raise ValueError("关联培训班不存在")

        self._update_training_exam(exam, data)
        self.db.commit()
        self.db.refresh(exam)

        detail = self.get_exam_detail(exam.id)
        if not detail:
            raise ValueError("考试不存在")
        return detail

    def update_admission_exam(self, exam_id: int, data: AdmissionExamUpdate) -> AdmissionExamResponse:
        """更新独立准入考试"""
        exam = self.db.query(AdmissionExam).options(
            *self._admission_exam_load_options(include_linked_trainings=False, include_records=False),
        ).filter(AdmissionExam.id == exam_id).first()
        if not exam:
            raise ValueError("准入考试不存在")

        self._update_admission_exam_entity(exam, data)
        self.db.commit()
        self.db.refresh(exam)

        detail = self.get_admission_exam_detail(exam.id)
        if not detail:
            raise ValueError("准入考试不存在")
        return detail

    def delete_exam(self, exam_id: int) -> None:
        """删除统一考试"""
        exam = self.db.query(Exam).options(
            *self._training_exam_load_options(include_paper=False, include_training=False),
        ).filter(Exam.id == exam_id).first()
        if not exam:
            raise ValueError("考试不存在")
        if exam.records:
            raise ValueError("考试已有作答记录，不能删除")
        if exam.participants:
            self.db.query(ExamParticipant).filter(ExamParticipant.exam_id == exam.id).delete(synchronize_session=False)

        self.db.delete(exam)
        self.db.commit()

    def delete_admission_exam(self, exam_id: int) -> None:
        """删除独立准入考试"""
        exam = self.db.query(AdmissionExam).options(
            *self._admission_exam_load_options(include_paper=False),
        ).filter(AdmissionExam.id == exam_id).first()
        if not exam:
            raise ValueError("准入考试不存在")
        if exam.records:
            raise ValueError("准入考试已有作答记录，不能删除")
        if exam.linked_trainings:
            raise ValueError("准入考试已关联培训班，不能删除")

        self.db.delete(exam)
        self.db.commit()

    def get_exam_detail(
        self,
        exam_id: int,
        current_user_id: Optional[int] = None,
    ) -> Optional[ExamDetailResponse]:
        """获取统一考试详情"""
        exam = self.db.query(Exam).options(
            *self._training_exam_load_options(),
        ).filter(Exam.id == exam_id).first()
        if not exam:
            return None

        if self._refresh_exam_status(exam):
            self.db.commit()
        scope_context = self._get_scope_context(current_user_id)
        if scope_context and (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_TRAINING:
            if not self._can_view_training_exam_with_context(scope_context, exam):
                return None
        current_user = self._get_admission_scope_user(current_user_id)
        if current_user_id and self._is_student_user(current_user):
            if not self._can_access_exam_as_student(exam, current_user_id):
                return None

        return ExamDetailResponse(
            **self._to_exam_response(exam, current_user_id).model_dump(),
            questions=self._build_snapshot_responses(
                exam.paper,
                include_sensitive=self._can_view_exam_sensitive_fields(current_user_id),
            ),
        )

    def get_admission_exam_detail(
        self,
        exam_id: int,
        current_user_id: Optional[int] = None,
    ) -> Optional[AdmissionExamDetailResponse]:
        """获取准入考试详情"""
        exam = self.db.query(AdmissionExam).options(
            *self._admission_exam_load_options(),
        ).filter(AdmissionExam.id == exam_id).first()
        if exam:
            if self._refresh_exam_status(exam):
                self.db.commit()

            current_user = self._get_admission_scope_user(current_user_id)
            can_manage_all = bool(current_user_id and self._can_manage_admission_exam(current_user_id))
            if current_user_id and not can_manage_all:
                if not current_user or not self._can_view_admission_exam(exam, current_user):
                    return None

            return AdmissionExamDetailResponse(
                **self._to_admission_exam_response(exam, current_user_id, current_user).model_dump(),
                questions=self._build_snapshot_responses(
                    exam.paper,
                    include_sensitive=self._can_view_exam_sensitive_fields(current_user_id),
                ),
            )

        unified_exam = self.db.query(Exam).options(
            *self._training_exam_load_options(),
        ).filter(
            Exam.id == exam_id,
            Exam.scene == EXAM_SCENE_STANDALONE,
            Exam.purpose == "admission",
        ).first()
        if not unified_exam:
            return None

        if self._refresh_exam_status(unified_exam):
            self.db.commit()
        if current_user_id and self._is_student_user(self._get_admission_scope_user(current_user_id)):
            if not self._can_access_exam_as_student(unified_exam, current_user_id):
                return None

        return AdmissionExamDetailResponse(
            **self._to_unified_exam_as_admission_response(unified_exam, current_user_id).model_dump(),
            questions=self._build_snapshot_responses(
                unified_exam.paper,
                include_sensitive=self._can_view_exam_sensitive_fields(current_user_id),
            ),
        )

    def submit_exam(self, exam_id: int, user_id: int, data: ExamSubmit) -> ExamRecordResponse:
        """提交统一考试"""
        exam = self.db.query(Exam).options(
            *self._training_exam_load_options(include_records=False),
        ).filter(Exam.id == exam_id).first()
        if not exam:
            raise ValueError("考试不存在")

        if self._refresh_exam_status(exam):
            self.db.commit()
        if exam.status != "active":
            raise ValueError("当前考试未开放作答")

        attempt_count = self._count_submitted_attempts(ExamRecord, ExamRecord.exam_id, exam.id, user_id)
        if not self._can_join_exam(exam, user_id, attempt_count):
            raise ValueError("当前用户无权参加该考试")

        record = self._build_exam_record(
            exam=exam,
            user_id=user_id,
            data=data,
            record_class=ExamRecord,
            exam_field_name="exam_id",
            attempt_count=attempt_count,
        )
        self.db.add(record)
        self._update_user_exam_stats(user_id, record.score or 0)
        self.db.commit()

        saved = self.db.query(ExamRecord).options(
            joinedload(ExamRecord.user),
            joinedload(ExamRecord.exam),
        ).filter(ExamRecord.id == record.id).first()
        logger.info("用户%s提交统一考试%s，得分=%s", user_id, exam_id, saved.score if saved else 0)
        self._mark_exam_participant_submitted(exam.id, user_id)
        return self._to_training_record_response(saved or record, exam)

    def submit_admission_exam(self, exam_id: int, user_id: int, data: ExamSubmit) -> AdmissionExamRecordResponse:
        """提交准入考试"""
        exam = self.db.query(AdmissionExam).options(
            *self._admission_exam_load_options(include_linked_trainings=False, include_records=False),
        ).filter(AdmissionExam.id == exam_id).first()
        if not exam:
            base = self.submit_exam(exam_id, user_id, data)
            return AdmissionExamRecordResponse(
                id=base.id,
                exam_id=base.exam_id,
                paper_id=base.paper_id,
                exam_title=base.exam_title,
                user_id=base.user_id,
                user_name=base.user_name,
                user_nickname=base.user_nickname,
                attempt_no=base.attempt_no,
                status=base.status,
                score=base.score,
                result=base.result,
                grade=base.grade,
                passing_score=base.passing_score,
                start_time=base.start_time,
                end_time=base.end_time,
                duration=base.duration,
                correct_count=base.correct_count,
                wrong_count=base.wrong_count,
                wrong_questions=base.wrong_questions,
                wrong_question_details=base.wrong_question_details,
                question_details=base.question_details,
                dimension_scores=base.dimension_scores,
            )

        if self._refresh_exam_status(exam):
            self.db.commit()
        if exam.status != "active":
            raise ValueError("当前准入考试未开放作答")

        attempt_count = self._count_submitted_attempts(
            AdmissionExamRecord,
            AdmissionExamRecord.admission_exam_id,
            exam.id,
            user_id,
        )
        current_user = self._get_admission_scope_user(user_id)
        if not self._can_join_admission_exam(exam, user_id, attempt_count, current_user):
            raise ValueError("当前用户不在准入考试适用范围内或已达到最大作答次数")

        record = self._build_exam_record(
            exam=exam,
            user_id=user_id,
            data=data,
            record_class=AdmissionExamRecord,
            exam_field_name="admission_exam_id",
            attempt_count=attempt_count,
        )
        self.db.add(record)
        self._update_user_exam_stats(user_id, record.score or 0)
        self.db.commit()

        saved = self.db.query(AdmissionExamRecord).options(
            joinedload(AdmissionExamRecord.user),
            joinedload(AdmissionExamRecord.admission_exam),
        ).filter(AdmissionExamRecord.id == record.id).first()
        logger.info("用户%s提交准入考试%s，得分=%s", user_id, exam_id, saved.score if saved else 0)
        return self._to_admission_record_response(saved or record, exam)

    def get_exam_result(self, exam_id: int, user_id: int) -> Optional[ExamRecordResponse]:
        """获取统一考试结果"""
        record = self.db.query(ExamRecord).options(
            joinedload(ExamRecord.user),
            joinedload(ExamRecord.exam),
        ).filter(
            ExamRecord.exam_id == exam_id,
            ExamRecord.user_id == user_id,
        ).order_by(ExamRecord.end_time.desc(), ExamRecord.id.desc()).first()
        if not record:
            return None
        return self._to_training_record_response(record, record.exam)

    def get_admission_exam_result(self, exam_id: int, user_id: int) -> Optional[AdmissionExamRecordResponse]:
        """获取准入考试结果"""
        record = self.db.query(AdmissionExamRecord).options(
            joinedload(AdmissionExamRecord.user),
            joinedload(AdmissionExamRecord.admission_exam),
        ).filter(
            AdmissionExamRecord.admission_exam_id == exam_id,
            AdmissionExamRecord.user_id == user_id,
        ).order_by(AdmissionExamRecord.end_time.desc(), AdmissionExamRecord.id.desc()).first()
        if record:
            return self._to_admission_record_response(record, record.admission_exam)
        unified_record = self.db.query(ExamRecord).options(
            joinedload(ExamRecord.user),
            joinedload(ExamRecord.exam),
        ).filter(
            ExamRecord.exam_id == exam_id,
            ExamRecord.user_id == user_id,
        ).order_by(ExamRecord.end_time.desc(), ExamRecord.id.desc()).first()
        if not unified_record:
            return None
        return self._to_exam_record_as_admission_response(unified_record, unified_record.exam)

    def get_exam_scores(self, exam_id: int, page: int = 1, size: int = 10) -> PaginatedResponse[ExamRecordResponse]:
        """获取统一考试成绩列表"""
        query = self.db.query(ExamRecord).options(
            joinedload(ExamRecord.user),
            joinedload(ExamRecord.exam),
        ).filter(
            ExamRecord.exam_id == exam_id,
        ).order_by(ExamRecord.score.desc(), ExamRecord.id.asc())
        total = query.count()
        rows = query.all() if size == -1 else query.offset((page - 1) * size).limit(size).all()
        return PaginatedResponse(
            page=page,
            size=size if size != -1 else total,
            total=total,
            items=[self._to_training_record_response(row, row.exam) for row in rows],
        )

    def get_admission_exam_scores(
        self,
        exam_id: int,
        page: int = 1,
        size: int = 10,
    ) -> PaginatedResponse[AdmissionExamRecordResponse]:
        """获取准入考试成绩列表"""
        query = self.db.query(AdmissionExamRecord).options(
            joinedload(AdmissionExamRecord.user),
            joinedload(AdmissionExamRecord.admission_exam),
        ).filter(
            AdmissionExamRecord.admission_exam_id == exam_id,
        ).order_by(AdmissionExamRecord.score.desc(), AdmissionExamRecord.id.asc())
        total = query.count()
        rows = query.all() if size == -1 else query.offset((page - 1) * size).limit(size).all()
        if total == 0:
            unified_query = self.db.query(ExamRecord).options(
                joinedload(ExamRecord.user),
                joinedload(ExamRecord.exam),
            ).filter(
                ExamRecord.exam_id == exam_id,
            ).order_by(ExamRecord.score.desc(), ExamRecord.id.asc())
            total = unified_query.count()
            unified_rows = unified_query.all() if size == -1 else unified_query.offset((page - 1) * size).limit(size).all()
            return PaginatedResponse(
                page=page,
                size=size if size != -1 else total,
                total=total,
                items=[self._to_exam_record_as_admission_response(row, row.exam) for row in unified_rows],
            )
        return PaginatedResponse(
            page=page,
            size=size if size != -1 else total,
            total=total,
            items=[self._to_admission_record_response(row, row.admission_exam) for row in rows],
        )

    def get_exam_analysis(self, exam_id: int) -> Dict[str, Any]:
        """获取统一考试分析"""
        exam = self.db.query(Exam).filter(Exam.id == exam_id).first()
        if not exam:
            raise ValueError("考试不存在")
        return self._build_analysis(
            self.db.query(ExamRecord).options(
                joinedload(ExamRecord.user).joinedload(User.departments),
            ).filter(
                ExamRecord.exam_id == exam_id,
                ExamRecord.status == "submitted",
            ).order_by(ExamRecord.score.desc()).all(),
            exam.passing_score or 60,
        )

    def get_exam_dashboard(self) -> Dict[str, Any]:
        """获取统一考试看板"""
        exams = self.db.query(Exam).options(
            joinedload(Exam.training),
            selectinload(Exam.participants).joinedload(ExamParticipant.user).selectinload(User.departments),
            selectinload(Exam.participants).joinedload(ExamParticipant.user).selectinload(User.police_types),
        ).all()
        legacy_admission_exams = self.db.query(AdmissionExam).all()

        total_exams = len(exams) + len(legacy_admission_exams)
        standalone_exams = len([item for item in exams if (item.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_STANDALONE]) + len(legacy_admission_exams)
        training_exams = len([item for item in exams if (item.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_TRAINING])
        active_exams = len([item for item in exams if (item.status or "") == "active"]) + len([item for item in legacy_admission_exams if (item.status or "") == "active"])
        upcoming_exams = len([item for item in exams if (item.status or "") == "upcoming"]) + len([item for item in legacy_admission_exams if (item.status or "") == "upcoming"])
        ended_exams = len([item for item in exams if (item.status or "") == "ended"]) + len([item for item in legacy_admission_exams if (item.status or "") == "ended"])

        new_records = self.db.query(ExamRecord).all()
        legacy_records = self.db.query(AdmissionExamRecord).all()
        all_scores = [int(item.score or 0) for item in [*new_records, *legacy_records] if item.status == "submitted"]
        pass_records = [item for item in [*new_records, *legacy_records] if item.status == "submitted" and item.result == "pass"]
        avg_score = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0
        pass_rate = round(len(pass_records) / len(all_scores) * 100, 1) if all_scores else 0

        purpose_counts: Dict[str, int] = {}
        police_type_counts: Dict[str, int] = {}
        department_counts: Dict[str, int] = {}
        recent_exam_items: List[Dict[str, Any]] = []

        for exam in exams:
            purpose = exam.purpose or "other"
            purpose_counts[purpose] = purpose_counts.get(purpose, 0) + 1
            for name in self._resolve_police_type_names_by_ids(self._normalize_scope_target_ids(exam.police_type_ids)):
                police_type_counts[name] = police_type_counts.get(name, 0) + 1
            for name in self._resolve_department_names_by_ids(self._normalize_scope_target_ids(exam.department_ids)):
                department_counts[name] = department_counts.get(name, 0) + 1
            if (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_STANDALONE:
                for participant in (exam.participants or []):
                    user = participant.user
                    for department in (user.departments or []) if user else []:
                        department_counts[department.name] = department_counts.get(department.name, 0) + 1
                    for police_type in (user.police_types or []) if user else []:
                        police_type_counts[police_type.name] = police_type_counts.get(police_type.name, 0) + 1
            recent_exam_items.append({
                "id": exam.id,
                "title": exam.title,
                "scene": exam.scene or EXAM_SCENE_TRAINING,
                "purpose": exam.purpose or "completion",
                "status": exam.status or "upcoming",
                "participant_count": self._count_exam_expected_participants(exam),
                "submitted_count": self._count_exam_submitted_users(exam),
                "created_at": exam.created_at.isoformat() if exam.created_at else None,
            })

        for exam in legacy_admission_exams:
            purpose_counts["admission"] = purpose_counts.get("admission", 0) + 1
            recent_exam_items.append({
                "id": exam.id,
                "title": exam.title,
                "scene": EXAM_SCENE_STANDALONE,
                "purpose": "admission",
                "status": exam.status or "upcoming",
                "participant_count": 0,
                "submitted_count": self.db.query(func.count(AdmissionExamRecord.id)).filter(
                    AdmissionExamRecord.admission_exam_id == exam.id,
                    AdmissionExamRecord.status == "submitted",
                ).scalar() or 0,
                "created_at": exam.created_at.isoformat() if exam.created_at else None,
            })

        trend_map: Dict[str, Dict[str, Any]] = {}
        for record in new_records:
            if not record.end_time:
                continue
            month_key = record.end_time.strftime("%Y-%m")
            item = trend_map.setdefault(month_key, {"month": month_key, "exam_count": 0, "participant_count": 0, "scores": []})
            item["exam_count"] += 1
            item["participant_count"] += 1
            item["scores"].append(int(record.score or 0))
        for record in legacy_records:
            if not record.end_time:
                continue
            month_key = record.end_time.strftime("%Y-%m")
            item = trend_map.setdefault(month_key, {"month": month_key, "exam_count": 0, "participant_count": 0, "scores": []})
            item["exam_count"] += 1
            item["participant_count"] += 1
            item["scores"].append(int(record.score or 0))

        trend = []
        for month_key in sorted(trend_map.keys()):
            item = trend_map[month_key]
            trend.append({
                "month": month_key,
                "exam_count": item["exam_count"],
                "participant_count": item["participant_count"],
                "avg_score": round(sum(item["scores"]) / len(item["scores"]), 1) if item["scores"] else 0,
            })

        recent_exam_items.sort(key=lambda item: item.get("created_at") or "", reverse=True)
        return {
            "kpi": {
                "total_exams": total_exams,
                "standalone_exams": standalone_exams,
                "training_exams": training_exams,
                "active_exams": active_exams,
                "upcoming_exams": upcoming_exams,
                "ended_exams": ended_exams,
                "avg_score": avg_score,
                "pass_rate": pass_rate,
            },
            "trend": trend,
            "purpose_distribution": [{"name": key, "value": value} for key, value in sorted(purpose_counts.items())],
            "police_type_distribution": [{"name": key, "value": value} for key, value in sorted(police_type_counts.items(), key=lambda item: item[1], reverse=True)],
            "department_ranking": [{"name": key, "value": value} for key, value in sorted(department_counts.items(), key=lambda item: item[1], reverse=True)[:10]],
            "recent_exams": recent_exam_items[:10],
        }

    def build_exam_participant_import_template(self) -> bytes:
        """生成独立考试名单导入模板"""
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise ValueError("openpyxl 未安装，无法生成模板") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "独立考试名单"
        ws.append(["姓名", "警号", "手机号", "身份证号", "部门", "警种"])
        ws.append(["张三", "10001", "13800000001", "110101199001010011", "治安支队", "治安"])
        ws.append(["李四", "10002", "13800000002", "110101199001010022", "刑侦支队", "刑侦"])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def preview_exam_participant_import(
        self,
        exam_id: int,
        file_bytes: bytes,
        file_name: str,
        actor_user_id: int,
    ) -> ExamParticipantImportPreviewResponse:
        """预检独立考试名单导入"""
        exam = self._get_standalone_exam_for_import(exam_id)
        rows = self._read_participant_excel_rows(file_bytes)
        if not rows:
            raise ValueError("Excel 文件为空，或缺少可识别表头")

        matched_rows: List[ExamParticipantImportRowResponse] = []
        created_rows: List[ExamParticipantImportRowResponse] = []
        failed_rows: List[ExamParticipantImportRowResponse] = []

        for row_no, row in rows:
            try:
                preview_row = self._build_participant_preview_row(row_no, row)
            except ValueError as exc:
                failed_rows.append(ExamParticipantImportRowResponse(
                    row_no=row_no,
                    name=str(row.get("name") or "").strip() or None,
                    police_id=str(row.get("police_id") or "").strip() or None,
                    phone=str(row.get("phone") or "").strip() or None,
                    id_card_number=str(row.get("id_card_number") or "").strip() or None,
                    status=PARTICIPANT_IMPORT_STATUS_FAILED,
                    reason=str(exc),
                ))
                continue
            if preview_row.status == PARTICIPANT_IMPORT_STATUS_MATCHED:
                matched_rows.append(preview_row)
            elif preview_row.status == PARTICIPANT_IMPORT_STATUS_CREATED:
                created_rows.append(preview_row)
            else:
                failed_rows.append(preview_row)

        batch = ExamParticipantImportBatch(
            exam_id=exam.id,
            file_name=file_name or "participants.xlsx",
            status="preview",
            summary={
                "matched_rows": [row.model_dump() for row in matched_rows],
                "created_rows": [row.model_dump() for row in created_rows],
                "total_rows": len(rows),
            },
            failure_rows=[row.model_dump() for row in failed_rows],
            created_by=actor_user_id,
        )
        self.db.add(batch)
        self.db.commit()
        self.db.refresh(batch)

        return ExamParticipantImportPreviewResponse(
            batch_id=batch.id,
            exam_id=exam.id,
            file_name=batch.file_name,
            summary={
                "total_rows": len(rows),
                "matched_count": len(matched_rows),
                "created_count": len(created_rows),
                "failed_count": len(failed_rows),
            },
            matched_rows=matched_rows,
            created_rows=created_rows,
            failed_rows=failed_rows,
        )

    def confirm_exam_participant_import(
        self,
        request: ExamParticipantImportConfirmRequest,
        actor_user_id: int,
    ) -> ExamParticipantImportPreviewResponse:
        """确认导入独立考试名单"""
        batch = self.db.query(ExamParticipantImportBatch).filter(
            ExamParticipantImportBatch.id == request.batch_id,
        ).first()
        if not batch:
            raise ValueError("导入批次不存在")
        exam = self._get_standalone_exam_for_import(batch.exam_id)

        matched_rows = [
            ExamParticipantImportRowResponse.model_validate(item)
            for item in ((batch.summary or {}).get("matched_rows") or [])
        ]
        created_rows = [
            ExamParticipantImportRowResponse.model_validate(item)
            for item in ((batch.summary or {}).get("created_rows") or [])
        ]
        failed_rows = [
            ExamParticipantImportRowResponse.model_validate(item)
            for item in (batch.failure_rows or [])
        ]

        created_output_rows: List[ExamParticipantImportRowResponse] = []
        for row in matched_rows:
            user = self.db.query(User).filter(User.id == row.user_id).first()
            if not user:
                failed_rows.append(ExamParticipantImportRowResponse(
                    **row.model_dump(),
                    status=PARTICIPANT_IMPORT_STATUS_FAILED,
                    reason="匹配到的用户已不存在",
                ))
                continue
            self._upsert_exam_participant(
                exam=exam,
                user=user,
                batch=batch,
                row=row,
                generated_password=None,
                match_status=PARTICIPANT_IMPORT_STATUS_MATCHED,
            )

        for row in created_rows:
            try:
                user, generated_password = self._create_participant_user_from_preview_row(row)
                created_output_row = ExamParticipantImportRowResponse(
                    **row.model_dump(),
                    user_id=user.id,
                    generated_password=generated_password,
                )
                created_output_rows.append(created_output_row)
                self._upsert_exam_participant(
                    exam=exam,
                    user=user,
                    batch=batch,
                    row=created_output_row,
                    generated_password=generated_password,
                    match_status=PARTICIPANT_IMPORT_STATUS_CREATED,
                )
            except ValueError as exc:
                failed_rows.append(ExamParticipantImportRowResponse(
                    **row.model_dump(),
                    status=PARTICIPANT_IMPORT_STATUS_FAILED,
                    reason=str(exc),
                ))

        batch.status = "confirmed"
        batch.summary = {
            "matched_rows": [row.model_dump() for row in matched_rows],
            "created_rows": [row.model_dump() for row in created_output_rows],
            "total_rows": (batch.summary or {}).get("total_rows", 0),
        }
        batch.failure_rows = [row.model_dump() for row in failed_rows]
        self.db.commit()

        return ExamParticipantImportPreviewResponse(
            batch_id=batch.id,
            exam_id=exam.id,
            file_name=batch.file_name,
            summary={
                "total_rows": (batch.summary or {}).get("total_rows", 0),
                "matched_count": len(matched_rows),
                "created_count": len(created_output_rows),
                "failed_count": len(failed_rows),
            },
            matched_rows=matched_rows,
            created_rows=created_output_rows,
            failed_rows=failed_rows,
        )

    def list_exam_participants(self, exam_id: int) -> List[ExamParticipantResponse]:
        """获取考试参试名单"""
        exam = self.db.query(Exam).options(
            selectinload(Exam.participants).joinedload(ExamParticipant.user).options(
                selectinload(User.departments),
                selectinload(User.police_types),
            ),
        ).filter(Exam.id == exam_id).first()
        if not exam:
            raise ValueError("考试不存在")
        return [self._to_exam_participant_response(item) for item in (exam.participants or [])]

    def export_exam_participant_import_result(self, batch_id: int) -> bytes:
        """导出导入结果"""
        batch = self.db.query(ExamParticipantImportBatch).filter(
            ExamParticipantImportBatch.id == batch_id,
        ).first()
        if not batch:
            raise ValueError("导入批次不存在")
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise ValueError("openpyxl 未安装，无法导出导入结果") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "导入结果"
        ws.append(["行号", "姓名", "警号", "手机号", "身份证号", "用户名", "状态", "原因", "初始密码"])

        rows = []
        rows.extend((batch.summary or {}).get("matched_rows") or [])
        rows.extend((batch.summary or {}).get("created_rows") or [])
        rows.extend(batch.failure_rows or [])
        for item in rows:
            ws.append([
                item.get("row_no"),
                item.get("name"),
                item.get("police_id"),
                item.get("phone"),
                item.get("id_card_number"),
                item.get("username"),
                item.get("status"),
                item.get("reason"),
                item.get("generated_password"),
            ])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def get_admission_exam_analysis(self, exam_id: int) -> Dict[str, Any]:
        """获取准入考试分析"""
        exam = self.db.query(AdmissionExam).filter(AdmissionExam.id == exam_id).first()
        if exam:
            return self._build_analysis(
                self.db.query(AdmissionExamRecord).options(
                    joinedload(AdmissionExamRecord.user).joinedload(User.departments),
                ).filter(
                    AdmissionExamRecord.admission_exam_id == exam_id,
                    AdmissionExamRecord.status == "submitted",
                ).order_by(AdmissionExamRecord.score.desc()).all(),
                exam.passing_score or 60,
            )
        unified_exam = self.db.query(Exam).filter(Exam.id == exam_id).first()
        if not unified_exam:
            raise ValueError("准入考试不存在")
        return self._build_analysis(
            self.db.query(ExamRecord).options(
                joinedload(ExamRecord.user).joinedload(User.departments),
            ).filter(
                ExamRecord.exam_id == exam_id,
                ExamRecord.status == "submitted",
            ).order_by(ExamRecord.score.desc()).all(),
            unified_exam.passing_score or 60,
        )

    def _build_analysis(self, records: List[Any], passing_score: int) -> Dict[str, Any]:
        students = []
        for record in records:
            user = record.user
            department_name = "未知单位"
            if user and user.departments:
                department_name = user.departments[0].name
            dimensions = record.dimension_scores or {}
            students.append({
                "id": record.id,
                "name": user.nickname if user and user.nickname else (user.username if user else "未知"),
                "policeId": user.police_id if user else "",
                "unit": department_name,
                "score": record.score or 0,
                "law": dimensions.get("law", 0),
                "enforce": dimensions.get("enforce", 0),
                "evidence": dimensions.get("evidence", 0),
                "physical": dimensions.get("physical", 0),
                "ethic": dimensions.get("ethic", 0),
                "time": f"{record.duration or 0}分钟",
                "pass": (record.score or 0) >= passing_score,
            })
        return {
            "students": students,
            "passing_score": passing_score,
        }

    def _paper_load_options(self) -> tuple:
        return (
            selectinload(ExamPaper.paper_questions).joinedload(
                ExamPaperQuestion.question
            ).options(
                joinedload(Question.police_type),
                selectinload(Question.knowledge_points).joinedload(KnowledgePoint.course),
            ),
            selectinload(ExamPaper.training_exams),
            selectinload(ExamPaper.admission_exams),
            joinedload(ExamPaper.creator),
            joinedload(ExamPaper.folder),
        )

    def _training_exam_load_options(
        self,
        include_paper: bool = True,
        include_training: bool = True,
        include_records: bool = True,
    ) -> tuple:
        options = []
        if include_paper:
            options.append(selectinload(Exam.paper).options(*self._paper_load_options()))
        if include_training:
            options.append(joinedload(Exam.training))
        if include_records:
            options.append(selectinload(Exam.records))
        options.append(selectinload(Exam.participants).joinedload(ExamParticipant.user))
        return tuple(options)

    def _admission_exam_load_options(
        self,
        include_paper: bool = True,
        include_linked_trainings: bool = True,
        include_records: bool = True,
    ) -> tuple:
        options = []
        if include_paper:
            options.append(selectinload(AdmissionExam.paper).options(*self._paper_load_options()))
        if include_linked_trainings:
            options.append(selectinload(AdmissionExam.linked_trainings))
        if include_records:
            options.append(selectinload(AdmissionExam.records))
        return tuple(options)

    def _get_scope_context(self, current_user_id: Optional[int]) -> Optional[DataScopeContext]:
        if not current_user_id:
            return None
        return build_data_scope_context(self.db, current_user_id)

    def _can_view_paper_with_context(
        self,
        scope_context: Optional[DataScopeContext],
        paper: Optional[ExamPaper],
    ) -> bool:
        if not paper:
            return False
        if scope_context is None:
            return True

        paper_questions = sorted(paper.paper_questions or [], key=lambda item: item.sort_order or 0)
        if not paper_questions:
            return True

        for paper_question in paper_questions:
            question = paper_question.question
            if question is None:
                if paper.created_by and paper.created_by == scope_context.user_id:
                    continue
                return False
            if not can_view_question_with_context(scope_context, question):
                return False
        return True

    def _can_view_training_exam_with_context(
        self,
        scope_context: Optional[DataScopeContext],
        exam: Optional[Exam],
    ) -> bool:
        if not exam:
            return False
        if scope_context is None:
            return True
        if not exam.training:
            # 无关联培训班的考试允许查看
            return True
        return can_view_training_with_context(scope_context, exam.training)

    def _get_paper_detail_entity(self, paper_id: int) -> Optional[ExamPaper]:
        return self.db.query(ExamPaper).options(*self._paper_load_options()).filter(ExamPaper.id == paper_id).first()

    def _get_selectable_paper(self, paper_id: int, current_user_id: Optional[int] = None) -> ExamPaper:
        paper = self._get_paper_detail_entity(paper_id)
        if not paper:
            raise ValueError("试卷不存在")
        scope_context = self._get_scope_context(current_user_id)
        if scope_context and not self._can_view_paper_with_context(scope_context, paper):
            raise ValueError("无权使用该试卷")
        if paper.status != "published":
            raise ValueError("只能选择已发布试卷")
        if not (paper.paper_questions or []):
            raise ValueError("试卷未配置题目")
        return paper

    def _sum_question_scores(self, questions: List[Question]) -> int:
        return sum(int(question.score or 0) for question in questions)

    def _paper_usage_count(self, paper: ExamPaper) -> int:
        return len(paper.training_exams or []) + len(paper.admission_exams or [])

    def _load_questions(self, question_ids: List[int]) -> List[Question]:
        if not question_ids:
            raise ValueError("请至少选择一道题目")
        questions = self.db.query(Question).options(
            selectinload(Question.knowledge_points)
        ).filter(Question.id.in_(question_ids)).all()
        question_map = {question.id: question for question in questions}
        ordered_questions = [question_map.get(question_id) for question_id in question_ids]
        if not all(ordered_questions):
            raise ValueError("存在无效题目，无法创建试卷")
        return ordered_questions

    def _update_training_exam(self, exam: Exam, data: ExamUpdate) -> None:
        update_data = data.model_dump(exclude_unset=True)
        paper_id = update_data.pop("paper_id", None)
        next_course_ids = update_data.pop("course_ids", exam.course_ids or [])
        next_department_ids = update_data.pop("department_ids", exam.department_ids or [])
        next_police_type_ids = update_data.pop("police_type_ids", exam.police_type_ids or [])
        if paper_id is not None and paper_id != exam.paper_id:
            raise ValueError("考试已发布，不能更换试卷")

        next_start_time = update_data.get("start_time", exam.start_time)
        next_end_time = update_data.get("end_time", exam.end_time)
        next_duration = self._resolve_exam_duration(update_data.get("duration"), exam.duration)
        next_passing_score = self._resolve_existing_exam_passing_score(
            exam.paper,
            update_data.get("passing_score"),
            exam.passing_score,
        )
        start_time, end_time = self._validate_exam_configuration(
            total_score=exam.paper.total_score,
            duration=next_duration,
            passing_score=next_passing_score,
            start_time=next_start_time,
            end_time=next_end_time,
        )

        for field, value in update_data.items():
            setattr(exam, field, value)

        exam.duration = next_duration
        exam.passing_score = next_passing_score
        exam.course_ids = self._resolve_selected_course_ids(next_course_ids)
        exam.department_ids = self._normalize_scope_target_ids(next_department_ids)
        exam.police_type_ids = self._normalize_scope_target_ids(next_police_type_ids)
        if (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_STANDALONE:
            exam.training_id = None
        exam.start_time = start_time
        exam.end_time = end_time
        exam.total_score = exam.paper.total_score or exam.total_score or 0
        if exam.status in {"active", "ended"} and not exam.published_at:
            exam.published_at = self._current_time(exam.start_time, exam.end_time, exam.published_at)

    def _update_admission_exam_entity(self, exam: AdmissionExam, data: AdmissionExamUpdate) -> None:
        update_data = data.model_dump(exclude_unset=True)
        paper_id = update_data.pop("paper_id", None)
        if paper_id is not None and paper_id != exam.paper_id:
            raise ValueError("准入考试已发布，不能更换试卷")

        next_scope_type = update_data.pop("scope_type", exam.scope_type or ADMISSION_SCOPE_ALL)
        next_scope_target_ids = update_data.pop("scope_target_ids", exam.scope_target_ids or [])
        next_course_ids = update_data.pop("course_ids", exam.course_ids or [])
        update_data.pop("scope", None)
        next_start_time = update_data.get("start_time", exam.start_time)
        next_end_time = update_data.get("end_time", exam.end_time)
        next_duration = self._resolve_exam_duration(update_data.get("duration"), exam.duration)
        next_passing_score = self._resolve_existing_exam_passing_score(
            exam.paper,
            update_data.get("passing_score"),
            exam.passing_score,
        )
        start_time, end_time = self._validate_exam_configuration(
            total_score=exam.paper.total_score,
            duration=next_duration,
            passing_score=next_passing_score,
            start_time=next_start_time,
            end_time=next_end_time,
        )

        for field, value in update_data.items():
            setattr(exam, field, value)

        scope_type, scope_target_ids, scope_summary = self._prepare_admission_scope(
            next_scope_type,
            next_scope_target_ids,
        )
        exam.scope_type = scope_type
        exam.scope_target_ids = scope_target_ids
        exam.scope = scope_summary
        exam.course_ids = self._resolve_selected_course_ids(next_course_ids)
        exam.duration = next_duration
        exam.passing_score = next_passing_score
        exam.start_time = start_time
        exam.end_time = end_time
        exam.total_score = exam.paper.total_score or exam.total_score or 0
        if exam.status in {"active", "ended"} and not exam.published_at:
            exam.published_at = self._current_time(exam.start_time, exam.end_time, exam.published_at)

    def _resolve_exam_duration(self, preferred: Optional[int], fallback: Optional[int]) -> int:
        value = preferred if preferred is not None else fallback
        return int(value or DEFAULT_EXAM_DURATION)

    def _resolve_create_exam_passing_score(self, paper: ExamPaper, preferred: Optional[int]) -> int:
        total_score = int(paper.total_score or 0)
        if total_score <= 0:
            raise ValueError("试卷满分必须大于0")
        if preferred is not None:
            return int(preferred)

        fallback = int(paper.passing_score or 0)
        if 0 < fallback <= total_score:
            return fallback
        return max(1, int(ceil(total_score * DEFAULT_PASSING_SCORE_RATIO)))

    def _resolve_existing_exam_passing_score(
        self,
        paper: ExamPaper,
        preferred: Optional[int],
        fallback: Optional[int],
    ) -> int:
        if preferred is not None:
            return int(preferred)
        if fallback is not None:
            return int(fallback)
        return self._resolve_create_exam_passing_score(paper, None)

    def _validate_exam_configuration(
        self,
        *,
        total_score: Optional[int],
        duration: int,
        passing_score: int,
        start_time: Optional[datetime],
        end_time: Optional[datetime],
    ) -> tuple[Optional[datetime], Optional[datetime]]:
        resolved_total_score = int(total_score or 0)
        if resolved_total_score <= 0:
            raise ValueError("试卷满分必须大于0")
        if duration < 10:
            raise ValueError("考试时长不能少于10分钟")
        if passing_score < 1:
            raise ValueError("及格分不能小于1分")
        if passing_score > resolved_total_score:
            raise ValueError(f"及格分不能超过试卷满分（{resolved_total_score}分）")

        now = self._current_time(start_time, end_time)
        normalized_start = self._normalize_datetime(start_time, now.tzinfo)
        normalized_end = self._normalize_datetime(end_time, now.tzinfo)
        if normalized_start and normalized_end:
            if normalized_end <= normalized_start:
                raise ValueError("考试结束时间必须晚于开始时间")

        return normalized_start, normalized_end

    def _can_manage_admission_exam(self, user_id: int) -> bool:
        return is_admin_user(self.db, user_id) or is_instructor_user(self.db, user_id)

    def _get_admission_scope_user(self, user_id: Optional[int]) -> Optional[User]:
        if not user_id:
            return None
        return self.db.query(User).options(
            joinedload(User.roles),
            joinedload(User.departments),
        ).filter(
            User.id == user_id,
            User.is_active == True,
        ).first()

    def _normalize_scope_target_ids(self, values: Any) -> List[int]:
        normalized: List[int] = []
        seen = set()
        for raw_item in values or []:
            try:
                item = int(raw_item)
            except (TypeError, ValueError):
                continue
            if item <= 0 or item in seen:
                continue
            seen.add(item)
            normalized.append(item)
        return normalized

    def _prepare_admission_scope(
        self,
        scope_type: Optional[str],
        scope_target_ids: Any,
    ) -> tuple[str, List[int], str]:
        resolved_scope_type = str(scope_type or ADMISSION_SCOPE_ALL).strip() or ADMISSION_SCOPE_ALL
        target_ids = self._normalize_scope_target_ids(scope_target_ids)

        if resolved_scope_type == ADMISSION_SCOPE_ALL:
            return ADMISSION_SCOPE_ALL, [], "全体学员"

        if not target_ids:
            raise ValueError("请至少选择一个适用范围目标")

        if resolved_scope_type == ADMISSION_SCOPE_USER:
            users = self.db.query(User).options(joinedload(User.roles)).filter(
                User.id.in_(target_ids),
                User.is_active == True,
            ).all()
            user_map = {
                user.id: user
                for user in users
                if self._is_student_user(user)
            }
            ordered_users = [user_map.get(item_id) for item_id in target_ids]
            if not all(ordered_users):
                raise ValueError("指定用户中包含无效学员")
            return (
                resolved_scope_type,
                target_ids,
                self._build_scope_summary("指定用户", [self._display_user_name(user) for user in ordered_users]),
            )

        if resolved_scope_type == ADMISSION_SCOPE_DEPARTMENT:
            departments = self.db.query(Department).filter(
                Department.id.in_(target_ids),
                Department.is_active == True,
            ).all()
            department_map = {department.id: department for department in departments}
            ordered_departments = [department_map.get(item_id) for item_id in target_ids]
            if not all(ordered_departments):
                raise ValueError("指定部门中包含无效部门")
            return (
                resolved_scope_type,
                target_ids,
                self._build_scope_summary("指定部门", [department.name for department in ordered_departments]),
            )

        if resolved_scope_type == ADMISSION_SCOPE_ROLE:
            roles = self.db.query(Role).filter(
                Role.id.in_(target_ids),
                Role.is_active == True,
            ).all()
            role_map = {role.id: role for role in roles}
            ordered_roles = [role_map.get(item_id) for item_id in target_ids]
            if not all(ordered_roles):
                raise ValueError("指定角色中包含无效角色")
            return (
                resolved_scope_type,
                target_ids,
                self._build_scope_summary("指定角色", [role.name for role in ordered_roles]),
            )

        raise ValueError("不支持的适用范围类型")

    def _build_scope_summary(self, label: str, names: List[str]) -> str:
        cleaned_names = [str(name).strip() for name in names if str(name or "").strip()]
        if not cleaned_names:
            return label
        if len(cleaned_names) <= 3:
            return f"{label}：{'、'.join(cleaned_names)}"
        return f"{label}：{'、'.join(cleaned_names[:3])} 等{len(cleaned_names)}项"

    def _display_user_name(self, user: Optional[User]) -> str:
        if not user:
            return ""
        return str(user.nickname or user.username or user.id)

    def _is_student_user(self, user: Optional[User]) -> bool:
        if not user:
            return False
        return any(str(role.code or "").strip() == "student" for role in (user.roles or []))

    def _can_view_admission_exam(self, exam: AdmissionExam, user: Optional[User]) -> bool:
        if not user or not user.is_active or not self._is_student_user(user):
            return False

        scope_type = exam.scope_type or ADMISSION_SCOPE_ALL
        target_ids = set(self._normalize_scope_target_ids(exam.scope_target_ids))
        if scope_type == ADMISSION_SCOPE_ALL:
            return True
        if not target_ids:
            return False
        if scope_type == ADMISSION_SCOPE_USER:
            return user.id in target_ids
        if scope_type == ADMISSION_SCOPE_DEPARTMENT:
            return any(department.id in target_ids for department in (user.departments or []))
        if scope_type == ADMISSION_SCOPE_ROLE:
            return any(role.id in target_ids for role in (user.roles or []))
        return False

    def _get_standalone_exam_for_import(self, exam_id: int) -> Exam:
        exam = self.db.query(Exam).options(*self._training_exam_load_options()).filter(Exam.id == exam_id).first()
        if not exam:
            raise ValueError("考试不存在")
        if (exam.scene or EXAM_SCENE_TRAINING) != EXAM_SCENE_STANDALONE:
            raise ValueError("仅独立考试支持名单导入")
        return exam

    def _read_participant_excel_rows(self, file_bytes: bytes) -> List[tuple[int, Dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("openpyxl 未安装，无法读取 Excel") from exc

        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header_aliases = {
            "name": {"姓名", "name", "昵称"},
            "police_id": {"警号", "police_id", "policeid"},
            "phone": {"手机号", "phone", "mobile"},
            "id_card_number": {"身份证号", "id_card_number", "idcard"},
            "department_names": {"部门", "部门名称", "department"},
            "police_type_names": {"警种", "警种名称", "police_type"},
        }
        header_map: Dict[int, str] = {}
        for index, raw_value in enumerate(rows[0]):
            text = str(raw_value or "").strip().lower()
            if not text:
                continue
            for target, aliases in header_aliases.items():
                if text in {alias.lower() for alias in aliases}:
                    header_map[index] = target
                    break
        if "name" not in header_map.values():
            raise ValueError("缺少姓名列")

        parsed_rows: List[tuple[int, Dict[str, Any]]] = []
        for row_no, values in enumerate(rows[1:], start=2):
            payload: Dict[str, Any] = {}
            empty = True
            for index, key in header_map.items():
                raw_value = values[index] if index < len(values) else None
                text = str(raw_value or "").strip()
                payload[key] = text or None
                if text:
                    empty = False
            if empty:
                continue
            parsed_rows.append((row_no, payload))
        return parsed_rows

    def _build_participant_preview_row(
        self,
        row_no: int,
        row: Dict[str, Any],
    ) -> ExamParticipantImportRowResponse:
        name = str(row.get("name") or "").strip()
        police_id = str(row.get("police_id") or "").strip() or None
        phone = str(row.get("phone") or "").strip() or None
        id_card_number = str(row.get("id_card_number") or "").strip() or None
        if not name:
            raise ValueError("姓名不能为空")

        user = self._find_user_for_import(police_id=police_id, phone=phone, id_card_number=id_card_number)
        if user:
            return ExamParticipantImportRowResponse(
                row_no=row_no,
                name=name,
                police_id=police_id,
                phone=phone,
                id_card_number=id_card_number,
                username=user.username,
                status=PARTICIPANT_IMPORT_STATUS_MATCHED,
                user_id=user.id,
            )

        generated_username = self._build_generated_username(police_id=police_id, phone=phone)
        if not generated_username:
            raise ValueError("未匹配到现有账号时，警号或手机号至少填写一个")
        return ExamParticipantImportRowResponse(
            row_no=row_no,
            name=name,
            police_id=police_id,
            phone=phone,
            id_card_number=id_card_number,
            username=generated_username,
            status=PARTICIPANT_IMPORT_STATUS_CREATED,
        )

    def _find_user_for_import(
        self,
        *,
        police_id: Optional[str] = None,
        phone: Optional[str] = None,
        id_card_number: Optional[str] = None,
    ) -> Optional[User]:
        if police_id:
            user = self.db.query(User).filter(User.police_id == police_id).first()
            if user:
                return user
        if phone:
            user = self.db.query(User).filter(User.phone == phone).first()
            if user:
                return user
        if id_card_number:
            user = self.db.query(User).filter(User.id_card_number == id_card_number).first()
            if user:
                return user
        return None

    def _build_generated_username(self, *, police_id: Optional[str], phone: Optional[str]) -> Optional[str]:
        candidate = str(police_id or phone or "").strip()
        return candidate or None

    def _create_participant_user_from_preview_row(self, row: ExamParticipantImportRowResponse) -> tuple[User, str]:
        username = self._build_generated_username(police_id=row.police_id, phone=row.phone)
        if not username:
            raise ValueError("缺少可生成账号的警号或手机号")
        if self.db.query(User).filter(User.username == username).first():
            raise ValueError(f"用户名 {username} 已存在")
        if row.phone and self.db.query(User).filter(User.phone == row.phone).first():
            raise ValueError(f"手机号 {row.phone} 已存在")
        if row.police_id and self.db.query(User).filter(User.police_id == row.police_id).first():
            raise ValueError(f"警号 {row.police_id} 已存在")
        if row.id_card_number and self.db.query(User).filter(User.id_card_number == row.id_card_number).first():
            raise ValueError("身份证号已存在")

        generated_password = self._generate_initial_password()
        student_role = self._ensure_student_role()
        user = User(
            username=username,
            password_hash=auth_service.get_password_hash(generated_password),
            nickname=row.name,
            police_id=row.police_id,
            phone=row.phone,
            id_card_number=row.id_card_number,
            is_active=True,
        )
        user.roles = [student_role]
        self.db.add(user)
        self.db.flush()
        return user, generated_password

    def _ensure_student_role(self) -> Role:
        role = self.db.query(Role).filter(Role.code == "student").first()
        if not role:
            raise ValueError("student 角色不存在")
        return role

    def _generate_initial_password(self, length: int = 10) -> str:
        alphabet = string.ascii_letters + string.digits
        return "".join(secrets.choice(alphabet) for _ in range(length))

    def _upsert_exam_participant(
        self,
        *,
        exam: Exam,
        user: User,
        batch: ExamParticipantImportBatch,
        row: ExamParticipantImportRowResponse,
        generated_password: Optional[str],
        match_status: str,
    ) -> ExamParticipant:
        participant = self.db.query(ExamParticipant).filter(
            ExamParticipant.exam_id == exam.id,
            ExamParticipant.user_id == user.id,
        ).first()
        if not participant:
            participant = ExamParticipant(
                exam_id=exam.id,
                user_id=user.id,
            )
            self.db.add(participant)
        participant.import_batch_id = batch.id
        participant.source_row_no = row.row_no
        participant.source_snapshot = row.model_dump()
        participant.match_status = match_status
        participant.generated_password = generated_password
        participant.participation_status = "assigned"
        return participant

    def _to_exam_participant_response(self, participant: ExamParticipant) -> ExamParticipantResponse:
        user = participant.user
        return ExamParticipantResponse(
            id=participant.id,
            exam_id=participant.exam_id,
            user_id=participant.user_id,
            user_name=user.username if user else None,
            user_nickname=user.nickname if user else None,
            police_id=user.police_id if user else None,
            phone=user.phone if user else None,
            departments=[item.name for item in (user.departments or []) if item.name] if user else [],
            police_types=[item.name for item in (user.police_types or []) if item.name] if user else [],
            match_status=participant.match_status or PARTICIPANT_IMPORT_STATUS_MATCHED,
            participation_status=participant.participation_status or "assigned",
            source_row_no=participant.source_row_no,
            created_at=participant.created_at,
        )

    def _replace_paper_questions(self, paper_id: int, question_ids: List[int], questions: List[Question]) -> None:
        question_map = {question.id: question for question in questions}
        self.db.query(ExamPaperQuestion).filter(
            ExamPaperQuestion.paper_id == paper_id,
        ).delete(synchronize_session=False)
        for index, question_id in enumerate(question_ids):
            question = question_map[question_id]
            self.db.add(ExamPaperQuestion(
                paper_id=paper_id,
                question_id=question_id,
                sort_order=index,
                question_type=question.type,
                content=question.content,
                options=question.options,
                answer=question.answer,
                explanation=question.explanation,
                score=question.score or 0,
                knowledge_points=self._resolve_question_knowledge_points(question),
            ))

    def _build_snapshot_responses(
        self,
        paper: Optional[ExamPaper],
        include_sensitive: bool = True,
    ) -> List[ExamQuestionSnapshotResponse]:
        if not paper:
            return []
        return [
            ExamQuestionSnapshotResponse(
                id=item.question_id,
                type=item.question_type or (item.question.type if item.question else "single"),
                content=item.content or (item.question.content if item.question else ""),
                options=item.options if item.options is not None else (item.question.options if item.question else None),
                answer=(
                    item.answer if item.answer is not None else (item.question.answer if item.question else None)
                ) if include_sensitive else None,
                explanation=(
                    item.explanation if item.explanation is not None else (item.question.explanation if item.question else None)
                ) if include_sensitive else None,
                score=int(item.score or (item.question.score if item.question else 0) or 0),
                knowledge_points=(
                    item.knowledge_points
                    if item.knowledge_points is not None
                    else self._resolve_question_knowledge_points(item.question)
                ),
            )
            for item in sorted(paper.paper_questions or [], key=lambda row: row.sort_order or 0)
        ]

    def _to_paper_response(self, paper: ExamPaper) -> ExamPaperResponse:
        linked_exam_count = len(paper.training_exams or [])
        linked_admission_exam_count = len(paper.admission_exams or [])
        paper_summary = self._extract_paper_dimension_summary(paper)

        return ExamPaperResponse(
            id=paper.id,
            title=paper.title,
            description=paper.description,
            duration=paper.duration or 60,
            total_score=paper.total_score or 0,
            passing_score=paper.passing_score or 60,
            type=paper.type or "formal",
            status=paper.status or "draft",
            folder_id=paper.folder_id,
            folder_name=paper.folder.name if paper.folder else None,
            published_at=paper.published_at,
            created_by=paper.created_by,
            creator_name=paper.creator.nickname if paper.creator else None,
            question_count=len(paper.paper_questions or []),
            usage_count=linked_exam_count + linked_admission_exam_count,
            linked_exam_count=linked_exam_count,
            linked_admission_exam_count=linked_admission_exam_count,
            knowledge_point_names=paper_summary["knowledge_point_names"],
            police_type_id=paper_summary["police_type_id"],
            police_type_name=paper_summary["police_type_name"],
            course_id=paper_summary["course_id"],
            course_name=paper_summary["course_name"],
            course_ids=paper_summary["course_ids"],
            course_names=paper_summary["course_names"],
            created_at=paper.created_at,
            updated_at=paper.updated_at,
        )

    def _to_paper_detail_response(self, paper: ExamPaper) -> ExamPaperDetailResponse:
        return ExamPaperDetailResponse(
            **self._to_paper_response(paper).model_dump(),
            questions=self._build_snapshot_responses(paper),
        )

    def _build_exam_record(
        self,
        exam: Any,
        user_id: int,
        data: ExamSubmit,
        record_class: Any,
        exam_field_name: str,
        attempt_count: int,
    ) -> Any:
        snapshots = self._build_snapshot_responses(exam.paper)
        if not snapshots:
            raise ValueError("考试未配置题目")

        score = 0
        correct_count = 0
        wrong_count = 0
        wrong_questions: List[int] = []
        wrong_details: List[ExamWrongQuestionResponse] = []
        dimension_totals = {key: 0 for key in DIMENSION_KEYS}
        dimension_gains = {key: 0 for key in DIMENSION_KEYS}

        for snapshot in snapshots:
            question_score = int(snapshot.score or 0)
            dimension = self._resolve_dimension(snapshot.knowledge_points)
            dimension_totals[dimension] += question_score

            user_answer = data.answers.get(str(snapshot.id))
            correct_answer = self._resolve_correct_answer(exam.paper, snapshot.id)
            if self._is_correct_answer(snapshot.type, user_answer, correct_answer):
                score += question_score
                correct_count += 1
                dimension_gains[dimension] += question_score
                continue

            wrong_count += 1
            wrong_questions.append(snapshot.id)
            wrong_details.append(ExamWrongQuestionResponse(
                question_id=snapshot.id,
                type=snapshot.type,
                content=snapshot.content,
                my_answer=user_answer,
                answer=correct_answer,
                explanation=snapshot.explanation,
                score=question_score,
            ))

        now = self._current_time(data.start_time)
        start_time = self._normalize_datetime(data.start_time, now.tzinfo)
        duration = 0
        if start_time:
            duration = max(0, int((now - start_time).total_seconds() / 60))
        dimension_scores = {
            key: int(round((dimension_gains[key] / dimension_totals[key]) * 100))
            if dimension_totals[key] > 0 else 0
            for key in DIMENSION_KEYS
        }
        payload = {
            exam_field_name: exam.id,
            "paper_id": exam.paper_id,
            "user_id": user_id,
            "attempt_no": attempt_count + 1,
            "status": "submitted",
            "score": score,
            "result": "pass" if score >= int(exam.passing_score or 60) else "fail",
            "grade": self._resolve_grade(score),
            "start_time": start_time,
            "end_time": now,
            "duration": duration,
            "answers": data.answers,
            "correct_count": correct_count,
            "wrong_count": wrong_count,
            "wrong_questions": wrong_questions,
            "wrong_question_details": [row.model_dump() for row in wrong_details],
            "dimension_scores": dimension_scores,
        }
        return record_class(**payload)

    def _resolve_correct_answer(self, paper: Optional[ExamPaper], question_id: int) -> Any:
        if not paper:
            return None
        for row in paper.paper_questions or []:
            if row.question_id != question_id:
                continue
            if row.answer is not None:
                return row.answer
            if row.question:
                return row.question.answer
        return None

    def _count_submitted_attempts(self, model: Any, field: Any, exam_id: int, user_id: int) -> int:
        return self.db.query(model).filter(
            field == exam_id,
            model.user_id == user_id,
            model.status == "submitted",
        ).count()

    def _to_training_exam_response(self, exam: Exam, current_user_id: Optional[int] = None) -> ExamResponse:
        attempt_count = 0
        latest_result = None
        can_join = None
        if current_user_id:
            latest_record = self.db.query(ExamRecord).filter(
                ExamRecord.exam_id == exam.id,
                ExamRecord.user_id == current_user_id,
            ).order_by(ExamRecord.end_time.desc(), ExamRecord.id.desc()).first()
            if latest_record:
                attempt_count = self._count_submitted_attempts(ExamRecord, ExamRecord.exam_id, exam.id, current_user_id)
                latest_result = latest_record.result
            can_join = self._can_join_exam(exam, current_user_id, attempt_count)
        paper_summary = self._extract_paper_dimension_summary(exam.paper, exam.course_ids)
        department_ids = self._normalize_scope_target_ids(exam.department_ids)
        police_type_ids = self._normalize_scope_target_ids(exam.police_type_ids)
        participant_count = self._count_exam_expected_participants(exam)
        submitted_count = self._count_exam_submitted_users(exam)
        absent_count = max(0, participant_count - submitted_count)
        scene = exam.scene or (EXAM_SCENE_TRAINING if exam.training_id else EXAM_SCENE_STANDALONE)

        return ExamResponse(
            id=exam.id,
            kind=scene,
            scene=scene,
            participant_mode=exam.participant_mode or EXAM_PARTICIPANT_MODE_TRAINING,
            paper_id=exam.paper_id,
            paper_title=exam.paper.title if exam.paper else None,
            paper_status=exam.paper.status if exam.paper else None,
            title=exam.title,
            description=exam.description,
            duration=exam.duration or 60,
            total_score=exam.total_score or 0,
            passing_score=exam.passing_score or 60,
            status=exam.status or "upcoming",
            type=exam.type or "formal",
            purpose=exam.purpose or "completion",
            training_id=exam.training_id,
            training_name=exam.training.name if exam.training else None,
            course_id=paper_summary["course_id"],
            course_name=paper_summary["course_name"],
            course_ids=paper_summary["course_ids"],
            course_names=paper_summary["course_names"],
            department_ids=department_ids,
            police_type_ids=police_type_ids,
            department_names=self._resolve_department_names_by_ids(department_ids),
            police_type_names=self._resolve_police_type_names_by_ids(police_type_ids),
            participant_summary=exam.participant_summary,
            max_attempts=exam.max_attempts or 1,
            allow_makeup=bool(exam.allow_makeup),
            start_time=exam.start_time,
            end_time=exam.end_time,
            created_by=exam.created_by,
            question_count=len(exam.paper.paper_questions or []) if exam.paper else 0,
            participant_count=participant_count,
            submitted_count=submitted_count,
            absent_count=absent_count,
            attempt_count=attempt_count,
            latest_result=latest_result,
            can_join=can_join,
            created_at=exam.created_at,
            updated_at=exam.updated_at,
        )

    def _to_admission_exam_response(
        self,
        exam: AdmissionExam,
        current_user_id: Optional[int] = None,
        current_user: Optional[User] = None,
    ) -> AdmissionExamResponse:
        attempt_count = 0
        latest_result = None
        can_join = None
        if current_user_id:
            latest_record = self.db.query(AdmissionExamRecord).filter(
                AdmissionExamRecord.admission_exam_id == exam.id,
                AdmissionExamRecord.user_id == current_user_id,
            ).order_by(AdmissionExamRecord.end_time.desc(), AdmissionExamRecord.id.desc()).first()
            if latest_record:
                attempt_count = self._count_submitted_attempts(
                    AdmissionExamRecord,
                    AdmissionExamRecord.admission_exam_id,
                    exam.id,
                    current_user_id,
                )
                latest_result = latest_record.result
            can_join = self._can_join_admission_exam(exam, current_user_id, attempt_count, current_user)
        paper_summary = self._extract_paper_dimension_summary(exam.paper, exam.course_ids)

        return AdmissionExamResponse(
            id=exam.id,
            paper_id=exam.paper_id,
            paper_title=exam.paper.title if exam.paper else None,
            paper_status=exam.paper.status if exam.paper else None,
            title=exam.title,
            description=exam.description,
            duration=exam.duration or 60,
            total_score=exam.total_score or 0,
            passing_score=exam.passing_score or 60,
            status=exam.status or "upcoming",
            type=exam.type or "formal",
            scope=exam.scope or "全体学员",
            scope_type=exam.scope_type or ADMISSION_SCOPE_ALL,
            scope_target_ids=self._normalize_scope_target_ids(exam.scope_target_ids),
            max_attempts=exam.max_attempts or 1,
            linked_training_count=len(exam.linked_trainings or []),
            course_id=paper_summary["course_id"],
            course_name=paper_summary["course_name"],
            course_ids=paper_summary["course_ids"],
            course_names=paper_summary["course_names"],
            attempt_count=attempt_count,
            latest_result=latest_result,
            can_join=can_join,
            created_by=exam.created_by,
            question_count=len(exam.paper.paper_questions or []) if exam.paper else 0,
            start_time=exam.start_time,
            end_time=exam.end_time,
            created_at=exam.created_at,
            updated_at=exam.updated_at,
        )

    def _to_unified_exam_as_admission_response(
        self,
        exam: Exam,
        current_user_id: Optional[int] = None,
    ) -> AdmissionExamResponse:
        exam_response = self._to_training_exam_response(exam, current_user_id)
        return AdmissionExamResponse(
            id=exam_response.id,
            paper_id=exam_response.paper_id,
            paper_title=exam_response.paper_title,
            paper_status=exam_response.paper_status,
            title=exam_response.title,
            description=exam_response.description,
            duration=exam_response.duration,
            total_score=exam_response.total_score,
            passing_score=exam_response.passing_score,
            status=exam_response.status,
            type=exam_response.type,
            scope=exam.participant_summary or "Excel 导入名单",
            scope_type=ADMISSION_SCOPE_ALL,
            scope_target_ids=[],
            max_attempts=exam_response.max_attempts,
            linked_training_count=0,
            course_id=exam_response.course_id,
            course_name=exam_response.course_name,
            course_ids=exam_response.course_ids,
            course_names=exam_response.course_names,
            attempt_count=exam_response.attempt_count,
            latest_result=exam_response.latest_result,
            can_join=exam_response.can_join,
            created_by=exam_response.created_by,
            question_count=exam_response.question_count,
            start_time=exam_response.start_time,
            end_time=exam_response.end_time,
            created_at=exam_response.created_at,
            updated_at=exam_response.updated_at,
        )

    def _to_training_record_response(self, record: ExamRecord, exam: Optional[Exam]) -> ExamRecordResponse:
        details = [
            ExamWrongQuestionResponse.model_validate(item)
            for item in (record.wrong_question_details or [])
        ]
        question_details = self._build_record_question_details(
            answers=record.answers or {},
            paper=exam.paper if exam else None,
            paper_id=record.paper_id,
        )
        user = record.user if getattr(record, "user", None) else None
        return ExamRecordResponse(
            id=record.id,
            exam_id=record.exam_id,
            kind=exam.scene if exam and exam.scene else EXAM_SCENE_TRAINING,
            scene=exam.scene if exam and exam.scene else EXAM_SCENE_TRAINING,
            paper_id=record.paper_id,
            exam_title=exam.title if exam else None,
            user_id=record.user_id,
            user_name=user.username if user else None,
            user_nickname=user.nickname if user else None,
            attempt_no=record.attempt_no or 1,
            status=record.status or "submitted",
            score=record.score or 0,
            result=record.result,
            grade=record.grade,
            passing_score=exam.passing_score if exam else None,
            start_time=record.start_time,
            end_time=record.end_time,
            duration=record.duration or 0,
            correct_count=record.correct_count or 0,
            wrong_count=record.wrong_count or 0,
            wrong_questions=record.wrong_questions or [],
            wrong_question_details=details,
            question_details=question_details,
            dimension_scores=record.dimension_scores or {},
        )

    def _to_admission_record_response(
        self,
        record: AdmissionExamRecord,
        exam: Optional[AdmissionExam],
    ) -> AdmissionExamRecordResponse:
        details = [
            ExamWrongQuestionResponse.model_validate(item)
            for item in (record.wrong_question_details or [])
        ]
        question_details = self._build_record_question_details(
            answers=record.answers or {},
            paper=exam.paper if exam else None,
            paper_id=record.paper_id,
        )
        user = record.user if getattr(record, "user", None) else None
        return AdmissionExamRecordResponse(
            id=record.id,
            exam_id=record.admission_exam_id,
            paper_id=record.paper_id,
            exam_title=exam.title if exam else None,
            user_id=record.user_id,
            user_name=user.username if user else None,
            user_nickname=user.nickname if user else None,
            attempt_no=record.attempt_no or 1,
            status=record.status or "submitted",
            score=record.score or 0,
            result=record.result,
            grade=record.grade,
            passing_score=exam.passing_score if exam else None,
            start_time=record.start_time,
            end_time=record.end_time,
            duration=record.duration or 0,
            correct_count=record.correct_count or 0,
            wrong_count=record.wrong_count or 0,
            wrong_questions=record.wrong_questions or [],
            wrong_question_details=details,
            question_details=question_details,
            dimension_scores=record.dimension_scores or {},
        )

    def _to_exam_record_as_admission_response(
        self,
        record: ExamRecord,
        exam: Optional[Exam],
    ) -> AdmissionExamRecordResponse:
        base = self._to_training_record_response(record, exam)
        return AdmissionExamRecordResponse(
            id=base.id,
            exam_id=base.exam_id,
            paper_id=base.paper_id,
            exam_title=base.exam_title,
            user_id=base.user_id,
            user_name=base.user_name,
            user_nickname=base.user_nickname,
            attempt_no=base.attempt_no,
            status=base.status,
            score=base.score,
            result=base.result,
            grade=base.grade,
            passing_score=base.passing_score,
            start_time=base.start_time,
            end_time=base.end_time,
            duration=base.duration,
            correct_count=base.correct_count,
            wrong_count=base.wrong_count,
            wrong_questions=base.wrong_questions,
            wrong_question_details=base.wrong_question_details,
            question_details=base.question_details,
            dimension_scores=base.dimension_scores,
        )

    def _build_record_question_details(
        self,
        answers: Dict[str, Any],
        paper: Optional[ExamPaper],
        paper_id: Optional[int],
    ) -> List[ExamQuestionAnswerDetailResponse]:
        target_paper = paper or self._get_paper_snapshot_entity(paper_id)
        if not target_paper:
            return []

        snapshots = self._build_snapshot_responses(target_paper)
        details: List[ExamQuestionAnswerDetailResponse] = []
        for snapshot in snapshots:
            my_answer = answers.get(str(snapshot.id))
            if my_answer is None:
                my_answer = answers.get(snapshot.id)
            correct_answer = self._resolve_correct_answer(target_paper, snapshot.id)
            details.append(
                ExamQuestionAnswerDetailResponse(
                    question_id=snapshot.id,
                    type=snapshot.type,
                    content=snapshot.content,
                    my_answer=my_answer,
                    answer=correct_answer,
                    is_correct=self._is_correct_answer(snapshot.type, my_answer, correct_answer),
                    explanation=snapshot.explanation,
                    score=int(snapshot.score or 0),
                )
            )
        return details

    def _get_paper_snapshot_entity(self, paper_id: Optional[int]) -> Optional[ExamPaper]:
        if not paper_id:
            return None
        return self.db.query(ExamPaper).options(
            selectinload(ExamPaper.paper_questions).joinedload(ExamPaperQuestion.question),
        ).filter(ExamPaper.id == paper_id).first()

    @staticmethod
    def _is_aware_datetime(value: Optional[datetime]) -> bool:
        return bool(value and value.tzinfo and value.tzinfo.utcoffset(value) is not None)

    @classmethod
    def _current_time(cls, *references: Optional[datetime]) -> datetime:
        for reference in references:
            if cls._is_aware_datetime(reference):
                return datetime.now(reference.tzinfo)
        return datetime.now()

    @classmethod
    def _normalize_datetime(cls, value: Optional[datetime], reference_tzinfo: Any) -> Optional[datetime]:
        if value is None:
            return None
        if cls._is_aware_datetime(value) or reference_tzinfo is None:
            return value
        return value.replace(tzinfo=reference_tzinfo)

    def _refresh_exam_status(self, exam: Any) -> bool:
        now = self._current_time(exam.start_time, exam.end_time)
        start_time = self._normalize_datetime(exam.start_time, now.tzinfo)
        end_time = self._normalize_datetime(exam.end_time, now.tzinfo)

        # 没有时间信息时，保持不变（由人工设置状态）
        if not start_time and not end_time:
            return False

        next_status = "upcoming"
        if start_time and end_time:
            if now < start_time:
                next_status = "upcoming"
            elif start_time <= now <= end_time:
                next_status = "active"
            else:
                next_status = "ended"
        elif end_time:
            next_status = "ended" if now > end_time else "upcoming"
        elif start_time:
            next_status = "active" if now >= start_time else "upcoming"

        if next_status != exam.status:
            exam.status = next_status
            return True
        return False

    def _can_join_training_exam(self, exam: Exam, user_id: int, attempt_count: int) -> bool:
        if exam.status != "active":
            return False
        if attempt_count >= int(exam.max_attempts or 1):
            return False
        enrollment = self.db.query(Enrollment.id).filter(
            Enrollment.training_id == exam.training_id,
            Enrollment.user_id == user_id,
            Enrollment.status == "approved",
        ).first()
        return enrollment is not None

    def _can_join_standalone_exam(self, exam: Exam, user_id: int, attempt_count: int) -> bool:
        if exam.status != "active":
            return False
        if attempt_count >= int(exam.max_attempts or 1):
            return False
        participant = self.db.query(ExamParticipant.id).filter(
            ExamParticipant.exam_id == exam.id,
            ExamParticipant.user_id == user_id,
        ).first()
        return participant is not None

    def _can_join_exam(self, exam: Exam, user_id: int, attempt_count: int) -> bool:
        if (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_STANDALONE:
            return self._can_join_standalone_exam(exam, user_id, attempt_count)
        return self._can_join_training_exam(exam, user_id, attempt_count)

    def _can_access_training_exam_as_student(self, exam: Optional[Exam], user_id: int) -> bool:
        if not exam or not user_id:
            return False
        enrollment = self.db.query(Enrollment.id).filter(
            Enrollment.training_id == exam.training_id,
            Enrollment.user_id == user_id,
            Enrollment.status == "approved",
        ).first()
        if enrollment is not None:
            return True
        record = self.db.query(ExamRecord.id).filter(
            ExamRecord.exam_id == exam.id,
            ExamRecord.user_id == user_id,
            ExamRecord.status == "submitted",
        ).first()
        return record is not None

    def _can_access_standalone_exam_as_student(self, exam: Optional[Exam], user_id: int) -> bool:
        if not exam or not user_id:
            return False
        participant = self.db.query(ExamParticipant.id).filter(
            ExamParticipant.exam_id == exam.id,
            ExamParticipant.user_id == user_id,
        ).first()
        if participant is not None:
            return True
        record = self.db.query(ExamRecord.id).filter(
            ExamRecord.exam_id == exam.id,
            ExamRecord.user_id == user_id,
            ExamRecord.status == "submitted",
        ).first()
        return record is not None

    def _can_access_exam_as_student(self, exam: Optional[Exam], user_id: int) -> bool:
        if not exam:
            return False
        if (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_STANDALONE:
            return self._can_access_standalone_exam_as_student(exam, user_id)
        return self._can_access_training_exam_as_student(exam, user_id)

    def _mark_exam_participant_submitted(self, exam_id: int, user_id: int) -> None:
        participant = self.db.query(ExamParticipant).filter(
            ExamParticipant.exam_id == exam_id,
            ExamParticipant.user_id == user_id,
        ).first()
        if participant:
            participant.participation_status = "submitted"

    def _count_exam_expected_participants(self, exam: Exam) -> int:
        if (exam.scene or EXAM_SCENE_TRAINING) == EXAM_SCENE_STANDALONE:
            return len(exam.participants or [])
        if not exam.training_id:
            return 0
        return self.db.query(Enrollment.id).filter(
            Enrollment.training_id == exam.training_id,
            Enrollment.status == "approved",
        ).count()

    def _count_exam_submitted_users(self, exam: Exam) -> int:
        return self.db.query(func.count(func.distinct(ExamRecord.user_id))).filter(
            ExamRecord.exam_id == exam.id,
            ExamRecord.status == "submitted",
        ).scalar() or 0

    def _resolve_department_names_by_ids(self, department_ids: List[int]) -> List[str]:
        if not department_ids:
            return []
        items = self.db.query(Department).filter(Department.id.in_(department_ids)).all()
        item_map = {item.id: item.name for item in items}
        return [item_map[item_id] for item_id in department_ids if item_id in item_map]

    def _resolve_police_type_names_by_ids(self, police_type_ids: List[int]) -> List[str]:
        if not police_type_ids:
            return []
        items = self.db.query(PoliceType).filter(PoliceType.id.in_(police_type_ids)).all()
        item_map = {item.id: item.name for item in items}
        return [item_map[item_id] for item_id in police_type_ids if item_id in item_map]

    def _extract_paper_dimension_summary(
        self,
        paper: Optional[ExamPaper],
        explicit_course_ids: Optional[List[int]] = None,
    ) -> Dict[str, Any]:
        knowledge_point_names: List[str] = []
        seen_kp_names = set()
        police_type_ids = set()
        police_type_names = set()
        course_pairs = []
        seen_course_ids = set()

        for pq in (paper.paper_questions or []) if paper else []:
            if pq.knowledge_points:
                for kp_name in pq.knowledge_points:
                    if kp_name and kp_name not in seen_kp_names:
                        seen_kp_names.add(kp_name)
                        knowledge_point_names.append(kp_name)

            if pq.question and pq.question.police_type:
                police_type_ids.add(pq.question.police_type_id)
                police_type_names.add(pq.question.police_type.name)

            if pq.question and pq.question.knowledge_points:
                for kp in pq.question.knowledge_points:
                    if kp.course_id and kp.course and kp.course_id not in seen_course_ids:
                        seen_course_ids.add(kp.course_id)
                        course_pairs.append((kp.course_id, kp.course.title))

        police_type_id = None
        police_type_name = None
        if len(police_type_ids) == 1:
            police_type_id = next(iter(police_type_ids))
            police_type_name = next(iter(police_type_names))
        elif len(police_type_ids) > 1:
            police_type_id = 0
            police_type_name = "综合类"

        explicit_pairs = self._get_course_pairs(explicit_course_ids or [], strict=False)
        selected_course_pairs = explicit_pairs or course_pairs
        course_ids = [item[0] for item in selected_course_pairs]
        course_names = [item[1] for item in selected_course_pairs]
        course_id = None
        course_name = None
        if len(selected_course_pairs) == 1:
            course_id, course_name = selected_course_pairs[0]
        elif len(selected_course_pairs) > 1:
            course_id = 0
            course_name = "综合类"

        return {
            "knowledge_point_names": knowledge_point_names,
            "police_type_id": police_type_id,
            "police_type_name": police_type_name,
            "course_id": course_id,
            "course_name": course_name,
            "course_ids": course_ids,
            "course_names": course_names,
        }

    def _resolve_selected_course_ids(self, values: Optional[List[int]]) -> List[int]:
        return [item[0] for item in self._get_course_pairs(values or [], strict=True)]

    def _get_course_pairs(self, course_ids: List[int], strict: bool = False) -> List[tuple[int, str]]:
        normalized_ids = self._normalize_scope_target_ids(course_ids)
        if not normalized_ids:
            return []
        courses = self.db.query(Course).filter(Course.id.in_(normalized_ids)).all()
        course_map = {course.id: course for course in courses}
        ordered_courses = [course_map.get(course_id) for course_id in normalized_ids]
        if strict and not all(ordered_courses):
            raise ValueError("指定课程中包含无效课程")
        return [(course.id, course.title) for course in ordered_courses if course]

    def _can_join_admission_exam(
        self,
        exam: AdmissionExam,
        user_id: int,
        attempt_count: int,
        current_user: Optional[User] = None,
    ) -> bool:
        if exam.status != "active":
            return False
        if attempt_count >= int(exam.max_attempts or 1):
            return False
        user = current_user or self._get_admission_scope_user(user_id)
        return self._can_view_admission_exam(exam, user)

    def _is_correct_answer(self, question_type: str, user_answer: Any, correct_answer: Any) -> bool:
        if question_type == "multi":
            if not isinstance(user_answer, list) or not isinstance(correct_answer, list):
                return False
            return sorted(str(item) for item in user_answer) == sorted(str(item) for item in correct_answer)
        return self._normalize_scalar_answer(user_answer) == self._normalize_scalar_answer(correct_answer)

    def _normalize_scalar_answer(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip().upper()
        if text in {"TRUE", "T", "A", "YES", "Y"}:
            return "T"
        if text in {"FALSE", "F", "B", "NO", "N"}:
            return "F"
        return text

    @staticmethod
    def _resolve_question_knowledge_points(question: Optional[Question]) -> List[str]:
        if not question:
            return []
        knowledge_points = sorted(
            question.knowledge_points or [],
            key=lambda item: (item.name or "", item.id or 0),
        )
        return [item.name for item in knowledge_points if item.name]

    def _resolve_dimension(self, knowledge_points: Optional[List[str]]) -> str:
        if not knowledge_points:
            return "law"
        joined_text = " ".join(str(item or "") for item in knowledge_points)
        for dimension, keywords in DIMENSION_RULES.items():
            if any(keyword in joined_text for keyword in keywords):
                return dimension
        return "law"

    def _resolve_grade(self, score: int) -> str:
        if score >= 90:
            return "A"
        if score >= 80:
            return "B"
        if score >= 60:
            return "C"
        return "D"

    def _update_user_exam_stats(self, user_id: int, score: int) -> None:
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user:
            return
        previous_count = user.exam_count or 0
        previous_total = float(user.avg_score or 0) * previous_count
        user.exam_count = previous_count + 1
        user.avg_score = (previous_total + score) / user.exam_count

    def _can_view_exam_sensitive_fields(self, current_user_id: Optional[int]) -> bool:
        if not current_user_id:
            return False
        return is_admin_user(self.db, current_user_id) or is_instructor_user(self.db, current_user_id)

    def _paginate(self, items: List[Any], page: int, size: int) -> PaginatedResponse[Any]:
        total = len(items)
        page_items = items
        if size != -1:
            start = (page - 1) * size
            page_items = items[start:start + size]
        return PaginatedResponse(
            page=page,
            size=size if size != -1 else total,
            total=total,
            items=page_items,
        )
