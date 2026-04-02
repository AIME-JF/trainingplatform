"""
Excel批量导入服务（用户底库、培训学员/教官、课程/课次）
"""

import re
import time
import uuid
from io import BytesIO
from datetime import datetime, date, time as dt_time, timedelta
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from app.models import (
    User,
    Role,
    Department,
    PoliceType,
    Training,
    Enrollment,
    TrainingCourse,
)
from app.services.auth import auth_service


class BatchImportService:
    """批量导入服务"""

    DEFAULT_PASSWORD = "Police@123456"

    USER_FIELD_ALIASES: Dict[str, set[str]] = {
        "name": {
            "姓名",
            "名字",
            "昵称",
            "姓名姓名",
            "name",
            "nickname",
        },
        "username": {
            "账号",
            "用户名",
            "登录名",
            "用户账号",
            "username",
            "loginname",
            "login",
        },
        "password": {"密码", "登录密码", "password"},
        "police_id": {
            "警号",
            "警员编号",
            "编号",
            "policeid",
            "policeno",
            "badge",
            "badgeid",
        },
        "id_card_number": {
            "身份证号",
            "身份证",
            "身份证号码",
            "证件号码",
            "证件号",
            "idcard",
            "idcardnumber",
            "id_card",
        },
        "phone": {"手机号", "手机", "联系电话", "phone", "mobile", "tel"},
        "email": {"邮箱", "电子邮箱", "mail", "email"},
        "gender": {"性别", "gender", "sex"},
        "department_names": {"部门", "单位", "所属单位", "department", "dept", "organization"},
        "police_type_names": {"警种", "警种类别", "警务类型", "policetype", "police_type"},
        "role": {"角色", "账号角色", "用户角色", "role", "userrole"},
    }

    COURSE_FIELD_ALIASES: Dict[str, set[str]] = {
        "course_name": {"课程名称", "课程", "科目", "课名", "course", "course_name", "title"},
        "instructor_name": {"教官", "授课教官", "讲师", "教师", "instructor", "teacher", "lecturer"},
        "type": {"课程类型", "类型", "课型", "type"},
        "hours": {"课时", "学时", "时长", "hours", "hour", "duration"},
        "location": {"地点", "教室", "上课地点", "location", "classroom"},
    }

    SESSION_FIELD_ALIASES: Dict[str, set[str]] = {
        "course_name": {"课程名称", "课程", "科目", "课名", "course", "course_name", "title"},
        "instructor_name": {"教官", "授课教官", "讲师", "教师", "instructor", "teacher", "lecturer"},
        "type": {"课程类型", "类型", "课型", "type"},
        "date": {"日期", "上课日期", "课程日期", "date", "day"},
        "time_range": {"时间段", "时段", "上课时间", "time", "time_range", "timerange"},
        "time_start": {"开始时间", "开课时间", "time_start", "start", "starttime"},
        "time_end": {"结束时间", "下课时间", "time_end", "end", "endtime"},
        "hours": {"课时", "学时", "时长", "hours", "hour", "duration"},
        "location": {"地点", "教室", "上课地点", "location", "classroom"},
    }

    def __init__(self, db: Session):
        self.db = db

    # ===== Public APIs =====

    def import_police_base(self, file_bytes: bytes, default_role_code: str = "student") -> Dict[str, Any]:
        rows = self._read_excel_rows(file_bytes, self.USER_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        summary = {
            "total_rows": 0,
            "success_rows": 0,
            "created_count": 0,
            "updated_count": 0,
            "matched_count": 0,
            "skipped_count": 0,
            "skipped_rows": [],
            "default_password": self.DEFAULT_PASSWORD,
        }

        for row_number, row in rows:
            summary["total_rows"] += 1
            try:
                with self.db.begin_nested():
                    result = self._upsert_user_from_row(
                        row=row,
                        default_role_code=default_role_code,
                        force_role_code=None,
                    )
                    if result["created"]:
                        summary["created_count"] += 1
                    elif result["updated"]:
                        summary["updated_count"] += 1
                        summary["matched_count"] += 1
                    else:
                        summary["matched_count"] += 1

                    summary["success_rows"] += 1
            except Exception as exc:
                summary["skipped_rows"].append({"row": row_number, "reason": str(exc)})

        self.db.commit()
        summary["skipped_count"] = len(summary["skipped_rows"])
        return summary

    def import_training_students(
        self,
        training_id: int,
        file_bytes: bytes,
        default_role_code: str = "student",
    ) -> Dict[str, Any]:
        training = self.db.query(Training).filter(Training.id == training_id).first()
        if not training:
            raise ValueError("培训班不存在")

        rows = self._read_excel_rows(file_bytes, self.USER_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        summary = {
            "training_id": training_id,
            "total_rows": 0,
            "success_rows": 0,
            "created_count": 0,
            "updated_count": 0,
            "matched_count": 0,
            "account_opened_count": 0,
            "enrollment_added": 0,
            "enrollment_approved": 0,
            "skipped_count": 0,
            "skipped_rows": [],
            "default_password": self.DEFAULT_PASSWORD,
        }

        for row_number, row in rows:
            summary["total_rows"] += 1
            try:
                with self.db.begin_nested():
                    result = self._upsert_user_from_row(
                        row=row,
                        default_role_code=default_role_code,
                        force_role_code="student",
                    )
                    user: User = result["user"]

                    if result["created"]:
                        summary["created_count"] += 1
                        summary["account_opened_count"] += 1
                    elif result["updated"]:
                        summary["updated_count"] += 1
                        summary["matched_count"] += 1
                    else:
                        summary["matched_count"] += 1

                    enrollment = self.db.query(Enrollment).filter(
                        Enrollment.training_id == training_id,
                        Enrollment.user_id == user.id,
                    ).first()

                    if not enrollment:
                        enrollment = Enrollment(
                            training_id=training_id,
                            user_id=user.id,
                            status="approved",
                            note="Excel批量导入学员",
                        )
                        self.db.add(enrollment)
                        summary["enrollment_added"] += 1
                    elif enrollment.status != "approved":
                        enrollment.status = "approved"
                        summary["enrollment_approved"] += 1

                    summary["success_rows"] += 1
            except Exception as exc:
                summary["skipped_rows"].append({"row": row_number, "reason": str(exc)})

        self.db.commit()
        summary["skipped_count"] = len(summary["skipped_rows"])
        return summary

    def import_training_instructors(
        self,
        training_id: int,
        file_bytes: bytes,
    ) -> Dict[str, Any]:
        training = self.db.query(Training).filter(Training.id == training_id).first()
        if not training:
            raise ValueError("培训班不存在")

        rows = self._read_excel_rows(file_bytes, self.USER_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        summary = {
            "training_id": training_id,
            "total_rows": 0,
            "success_rows": 0,
            "created_count": 0,
            "updated_count": 0,
            "matched_count": 0,
            "account_opened_count": 0,
            "assigned_instructor_id": training.instructor_id,
            "skipped_count": 0,
            "skipped_rows": [],
            "default_password": self.DEFAULT_PASSWORD,
        }

        first_imported_instructor_id: Optional[int] = None

        for row_number, row in rows:
            summary["total_rows"] += 1
            try:
                with self.db.begin_nested():
                    result = self._upsert_user_from_row(
                        row=row,
                        default_role_code="instructor",
                        force_role_code="instructor",
                    )
                    user: User = result["user"]

                    if first_imported_instructor_id is None:
                        first_imported_instructor_id = user.id

                    if result["created"]:
                        summary["created_count"] += 1
                        summary["account_opened_count"] += 1
                    elif result["updated"]:
                        summary["updated_count"] += 1
                        summary["matched_count"] += 1
                    else:
                        summary["matched_count"] += 1

                    summary["success_rows"] += 1
            except Exception as exc:
                summary["skipped_rows"].append({"row": row_number, "reason": str(exc)})

        if training.instructor_id is None and first_imported_instructor_id is not None:
            training.instructor_id = first_imported_instructor_id
            summary["assigned_instructor_id"] = first_imported_instructor_id

        self.db.commit()
        summary["skipped_count"] = len(summary["skipped_rows"])
        return summary

    def import_training_courses(
        self,
        training_id: int,
        file_bytes: bytes,
        commit: bool = True,
    ) -> Dict[str, Any]:
        training = self.db.query(Training).filter(Training.id == training_id).first()
        if not training:
            raise ValueError("培训班不存在")

        rows = self._read_excel_rows(file_bytes, self.COURSE_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        existing_courses = self._load_training_course_payloads(training_id)
        existing_keys = {self._course_identity(course) for course in existing_courses}
        imported_keys: set[str] = set()
        appended_courses: List[Dict[str, Any]] = []
        skipped_rows: List[Dict[str, Any]] = []

        for row_number, row in rows:
            try:
                course = self._build_course_entry(row)
                course_key = self._course_identity(course)
                if course_key in existing_keys or course_key in imported_keys:
                    skipped_rows.append({"row": row_number, "reason": "课程已存在，已跳过"})
                    continue
                imported_keys.add(course_key)
                appended_courses.append(course)
            except Exception as exc:
                skipped_rows.append({"row": row_number, "reason": str(exc)})

        if appended_courses:
            self._persist_training_courses(training_id, existing_courses + appended_courses)
            if commit:
                self.db.commit()
        elif commit:
            self.db.rollback()

        return {
            "training_id": training_id,
            "total_rows": len(rows),
            "added_course_count": len(appended_courses),
            "course_count": len(appended_courses),
            "skipped_count": len(skipped_rows),
            "skipped_rows": skipped_rows,
            "mode": "append",
        }

    def import_training_sessions(
        self,
        training_id: int,
        file_bytes: bytes,
        commit: bool = True,
    ) -> Dict[str, Any]:
        training = self.db.query(Training).filter(Training.id == training_id).first()
        if not training:
            raise ValueError("培训班不存在")

        rows = self._read_excel_rows(file_bytes, self.SESSION_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        existing_courses = self._load_training_course_payloads(training_id)
        skipped_rows: List[Dict[str, Any]] = []
        course_match_failures: List[Dict[str, Any]] = []
        added_session_count = 0

        for row_number, row in rows:
            session_seed: Optional[Dict[str, Any]] = None
            try:
                session_seed = self._build_session_entry(row)
                target_course = self._match_training_course(
                    existing_courses,
                    session_seed["course_name"],
                    session_seed["instructor_name"],
                    session_seed["course_type"],
                )
                if target_course is None:
                    reason = "未找到匹配课程，请先导入课程"
                    course_match_failures.append(self._build_course_match_failure(row_number, session_seed, reason))
                    skipped_rows.append({"row": row_number, "reason": reason})
                    continue
                if self._schedule_exists(target_course.get("schedules") or [], session_seed["schedule"]):
                    skipped_rows.append({"row": row_number, "reason": "课次已存在，已跳过"})
                    continue

                target_course.setdefault("schedules", []).append(session_seed["schedule"])
                target_course["schedules"].sort(
                    key=lambda item: (item.get("date") or "", item.get("time_range") or "", item.get("location") or "")
                )
                if not target_course.get("location") and session_seed["schedule"].get("location"):
                    target_course["location"] = session_seed["schedule"].get("location")
                added_session_count += 1
            except Exception as exc:
                reason = str(exc)
                if self._is_course_match_failure_reason(reason):
                    if session_seed:
                        course_match_failures.append(self._build_course_match_failure(row_number, session_seed, reason))
                skipped_rows.append({"row": row_number, "reason": reason})

        if added_session_count > 0:
            self._persist_training_courses(training_id, existing_courses)
            if commit:
                self.db.commit()
        elif commit:
            self.db.rollback()

        return {
            "training_id": training_id,
            "total_rows": len(rows),
            "added_session_count": added_session_count,
            "schedule_count": added_session_count,
            "skipped_count": len(skipped_rows),
            "skipped_rows": skipped_rows,
            "course_match_failure_count": len(course_match_failures),
            "course_match_failures": course_match_failures,
            "course_match_failure_summary": self._summarize_course_match_failures(course_match_failures),
            "mode": "append",
        }

    def import_training_schedule(
        self,
        training_id: int,
        file_bytes: bytes,
        replace_existing: bool = True,
        commit: bool = True,
    ) -> Dict[str, Any]:
        summary = self.import_training_sessions(
            training_id,
            file_bytes,
            commit=commit,
        )
        summary["replace_existing"] = False
        return summary

    # ===== Row Parsing =====

    def _read_excel_rows(
        self,
        file_bytes: bytes,
        field_aliases: Dict[str, set[str]],
    ) -> List[Tuple[int, Dict[str, Any]]]:
        if not file_bytes:
            return []

        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - dependency error
            raise ValueError("服务器缺少openpyxl依赖，无法解析Excel文件") from exc

        try:
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            raise ValueError("Excel文件解析失败，请确认文件为.xlsx格式") from exc

        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            return []

        alias_map = self._build_alias_map(field_aliases)
        header_row = raw_rows[0]
        header_keys: List[str] = []
        for index, raw_header in enumerate(header_row):
            normalized = self._normalize_header(raw_header)
            canonical = alias_map.get(normalized, normalized or f"column_{index + 1}")
            header_keys.append(canonical)

        results: List[Tuple[int, Dict[str, Any]]] = []
        for row_index, values in enumerate(raw_rows[1:], start=2):
            row: Dict[str, Any] = {}
            has_value = False
            for idx, raw_value in enumerate(values):
                value = self._normalize_cell(raw_value)
                if value in (None, ""):
                    continue
                key = header_keys[idx] if idx < len(header_keys) else f"column_{idx + 1}"
                row[key] = value
                has_value = True

            if has_value and not self._is_example_row(row):
                results.append((row_index, row))

        return results

    def _build_alias_map(self, field_aliases: Dict[str, set[str]]) -> Dict[str, str]:
        alias_map: Dict[str, str] = {}
        for canonical, aliases in field_aliases.items():
            alias_map[self._normalize_header(canonical)] = canonical
            for alias in aliases:
                alias_map[self._normalize_header(alias)] = canonical
        return alias_map

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().lower()
        text = re.sub(r"[\s\-_/:：()（）【】\[\]·]+", "", text)
        return text

    @staticmethod
    def _normalize_cell(value: Any) -> Any:
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def _is_example_row(row: Dict[str, Any]) -> bool:
        for value in row.values():
            text = str(value or "").strip()
            if text:
                return text.startswith("【示例】") or text.startswith("__EXAMPLE__")
        return False

    # ===== User Upsert =====

    def _upsert_user_from_row(
        self,
        row: Dict[str, Any],
        default_role_code: str,
        force_role_code: Optional[str],
    ) -> Dict[str, Any]:
        name = self._to_text(row.get("name"))
        username = self._to_text(row.get("username"))
        password = self._to_text(row.get("password"))
        police_id = self._to_text(row.get("police_id"))
        id_card_number = self._to_text(row.get("id_card_number"))
        phone = self._normalize_phone(row.get("phone"))
        email = self._to_text(row.get("email"))
        gender = self._normalize_gender(row.get("gender"))

        if not any([name, username, police_id, id_card_number, phone]):
            raise ValueError("缺少身份字段（姓名/身份证号/警号/账号/手机号）")

        role_code = force_role_code or self._parse_role_code(row.get("role"), default_role_code)
        roles = self._resolve_roles([role_code])
        departments = self._resolve_departments(self._split_multi_values(row.get("department_names")))
        police_types = self._resolve_police_types(self._split_multi_values(row.get("police_type_names")))

        user = self._find_user(police_id=police_id, username=username, phone=phone)
        if user:
            updated = False

            if name and name != (user.nickname or ""):
                user.nickname = name
                updated = True

            if gender and gender != (user.gender or ""):
                user.gender = gender
                updated = True

            if police_id and police_id != (user.police_id or "") and self._field_available("police_id", police_id, user.id):
                user.police_id = police_id
                updated = True

            if id_card_number and id_card_number != (user.id_card_number or "") and self._field_available("id_card_number", id_card_number, user.id):
                user.id_card_number = id_card_number
                updated = True

            if phone and phone != (user.phone or "") and self._field_available("phone", phone, user.id):
                user.phone = phone
                updated = True

            if email and email != (user.email or "") and self._field_available("email", email, user.id):
                user.email = email
                updated = True

            if user.is_active is False:
                user.is_active = True
                updated = True

            if self._merge_roles(user, roles):
                updated = True
            if self._merge_departments(user, departments):
                updated = True
            if self._merge_police_types(user, police_types):
                updated = True

            self.db.flush()
            return {"user": user, "created": False, "updated": updated}

        final_username = self._build_unique_username(username or police_id or phone or name or "user")
        nickname = name or final_username

        new_user = User(
            username=final_username,
            password_hash=auth_service.get_password_hash(password or self.DEFAULT_PASSWORD),
            nickname=nickname,
            gender=gender,
            phone=phone,
            email=email,
            police_id=police_id,
            id_card_number=id_card_number,
            is_active=True,
        )
        if roles:
            new_user.roles = roles
        if departments:
            new_user.departments = departments
        if police_types:
            new_user.police_types = police_types

        self.db.add(new_user)
        self.db.flush()
        return {"user": new_user, "created": True, "updated": True}

    def _find_user(self, police_id: Optional[str], username: Optional[str], phone: Optional[str]) -> Optional[User]:
        if police_id:
            user = self.db.query(User).filter(User.police_id == police_id).first()
            if user:
                return user
        if username:
            user = self.db.query(User).filter(User.username == username).first()
            if user:
                return user
        if phone:
            user = self.db.query(User).filter(User.phone == phone).first()
            if user:
                return user
        return None

    def _field_available(self, field_name: str, value: str, exclude_user_id: int) -> bool:
        if not value:
            return False
        field = getattr(User, field_name)
        conflict = self.db.query(User.id).filter(
            field == value,
            User.id != exclude_user_id,
        ).first()
        return conflict is None

    def _build_unique_username(self, seed: str) -> str:
        base = re.sub(r"[^a-zA-Z0-9_]+", "", str(seed or "").strip())
        if not base:
            base = "user"
        base = base[:40]
        if len(base) < 3:
            base = f"{base}usr"

        candidate = base
        index = 1
        while self.db.query(User.id).filter(User.username == candidate).first():
            suffix = f"_{index}"
            candidate = f"{base[:50 - len(suffix)]}{suffix}"
            index += 1
        return candidate

    def _merge_roles(self, user: User, roles: List[Role]) -> bool:
        changed = False
        existing = {r.id for r in (user.roles or [])}
        for role in roles:
            if role.id not in existing:
                user.roles.append(role)
                changed = True
        return changed

    def _merge_departments(self, user: User, departments: List[Department]) -> bool:
        if not departments:
            return False
        changed = False
        existing = {d.id for d in (user.departments or [])}
        for dept in departments:
            if dept.id not in existing:
                user.departments.append(dept)
                changed = True
        return changed

    def _merge_police_types(self, user: User, police_types: List[PoliceType]) -> bool:
        if not police_types:
            return False
        changed = False
        existing = {p.id for p in (user.police_types or [])}
        for police_type in police_types:
            if police_type.id not in existing:
                user.police_types.append(police_type)
                changed = True
        return changed

    def _resolve_roles(self, role_codes: List[str]) -> List[Role]:
        codes = [code for code in role_codes if code]
        if not codes:
            return []
        roles = self.db.query(Role).filter(Role.code.in_(codes)).all()
        role_map = {r.code: r for r in roles}
        missing = [code for code in codes if code not in role_map]
        if missing:
            raise ValueError(f"角色不存在: {','.join(missing)}")
        return [role_map[code] for code in codes]

    def _resolve_departments(self, names: List[str]) -> List[Department]:
        result: List[Department] = []
        for name in names:
            dept = self.db.query(Department).filter(Department.name == name).first()
            if not dept:
                dept = Department(
                    name=name,
                    code=self._generate_unique_code("dept"),
                    is_active=True,
                    inherit_sub_permissions=True,
                )
                self.db.add(dept)
                self.db.flush()
            result.append(dept)
        return result

    def _resolve_police_types(self, names: List[str]) -> List[PoliceType]:
        result: List[PoliceType] = []
        for name in names:
            police_type = self.db.query(PoliceType).filter(PoliceType.name == name).first()
            if not police_type:
                police_type = PoliceType(
                    name=name,
                    code=self._generate_unique_code("pt"),
                    is_active=True,
                )
                self.db.add(police_type)
                self.db.flush()
            result.append(police_type)
        return result

    def _generate_unique_code(self, prefix: str) -> str:
        while True:
            candidate = f"{prefix}_{int(time.time() * 1000) % 1000000000}_{uuid.uuid4().hex[:4]}"
            exists_dept = self.db.query(Department.id).filter(Department.code == candidate).first()
            exists_pt = self.db.query(PoliceType.id).filter(PoliceType.code == candidate).first()
            if not exists_dept and not exists_pt:
                return candidate

    def _parse_role_code(self, value: Any, default_code: str) -> str:
        text = self._to_text(value).lower() if self._to_text(value) else ""
        mapping = {
            "admin": "admin",
            "管理员": "admin",
            "系统管理员": "admin",
            "instructor": "instructor",
            "教官": "instructor",
            "讲师": "instructor",
            "教师": "instructor",
            "student": "student",
            "学员": "student",
            "民警": "student",
        }
        return mapping.get(text, default_code or "student")

    @staticmethod
    def _to_text(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        text = str(value).strip()
        return text or None

    @staticmethod
    def _normalize_phone(value: Any) -> Optional[str]:
        text = BatchImportService._to_text(value)
        if not text:
            return None
        cleaned = re.sub(r"\s+", "", text)
        return cleaned or None

    @staticmethod
    def _normalize_gender(value: Any) -> Optional[str]:
        text = BatchImportService._to_text(value)
        if not text:
            return None
        normalized = text.strip().lower()
        if normalized in {"男", "m", "male", "1"}:
            return "男"
        if normalized in {"女", "f", "female", "0"}:
            return "女"
        return text

    @staticmethod
    def _split_multi_values(value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            values = value
        else:
            values = re.split(r"[,，;；/、|]+", str(value))
        result = []
        seen = set()
        for item in values:
            text = str(item).strip()
            if not text:
                continue
            if text in seen:
                continue
            seen.add(text)
            result.append(text)
        return result

    # ===== Course / Session Import Helpers =====

    def _build_course_entry(self, row: Dict[str, Any]) -> Dict[str, Any]:
        course_name = self._to_text(row.get("course_name"))
        if not course_name:
            raise ValueError("缺少课程名称")

        instructor_name = self._to_text(row.get("instructor_name"))
        course_type = self._parse_optional_course_type(row.get("type"))
        hours = self._parse_float(row.get("hours"))
        raw_hours = row.get("hours")
        if raw_hours not in (None, "") and hours is None:
            raise ValueError("课时格式无效")
        if hours is not None and hours < 0:
            raise ValueError("课时不能小于 0")

        # 优先按名称+教官匹配课程资源
        course_id = self._match_course_resource(course_name, instructor_name)

        return {
            "course_id": course_id,
            "course_key": str(uuid.uuid4()),
            "name": course_name,
            "location": self._to_text(row.get("location")),
            "instructor": instructor_name,
            "primary_instructor_id": self._guess_instructor_id(instructor_name),
            "assistant_instructor_ids": [],
            "type": course_type,
            "hours": round(hours or 0, 2),
            "schedules": [],
        }

    def _match_course_resource(self, course_name: str, instructor_name: Optional[str] = None) -> Optional[int]:
        """按名称和教官匹配课程资源，返回 course_id 或 None"""
        from app.models.course import Course as CourseModel
        query = self.db.query(CourseModel).filter(CourseModel.title == course_name)
        candidates = query.all()
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0].id
        # 多个同名课程时，尝试用教官名进一步匹配
        if instructor_name:
            for c in candidates:
                inst = c.instructor
                if inst and (inst.nickname == instructor_name or inst.username == instructor_name):
                    return c.id
        # 无法精确匹配时返回第一个
        return candidates[0].id

    def _build_session_entry(self, row: Dict[str, Any]) -> Dict[str, Any]:
        course_name = self._to_text(row.get("course_name"))
        if not course_name:
            raise ValueError("缺少课程名称")

        instructor_name = self._to_text(row.get("instructor_name"))
        course_type = self._parse_optional_course_type(row.get("type"))
        parsed_date = self._parse_date(row.get("date"))
        if not parsed_date:
            raise ValueError("缺少有效上课日期")

        time_start = self._parse_time(row.get("time_start"))
        time_end = self._parse_time(row.get("time_end"))
        if not time_start or not time_end:
            range_start, range_end = self._parse_time_range(row.get("time_range"))
            time_start = time_start or range_start
            time_end = time_end or range_end
        if not time_start or not time_end:
            raise ValueError("缺少有效开始时间或结束时间")

        hours = self._parse_float(row.get("hours"))
        if hours is None or hours <= 0:
            hours = self._calc_hours(time_start, time_end)
        if hours <= 0:
            raise ValueError("课次时间无效，请检查开始和结束时间")

        return {
            "course_name": course_name,
            "instructor_name": instructor_name,
            "course_type": course_type,
            "schedule": {
                "session_id": str(uuid.uuid4()),
                "date": parsed_date,
                "time_range": f"{time_start}~{time_end}",
                "hours": round(hours, 2),
                "location": self._to_text(row.get("location")),
            },
        }

    def _load_training_course_payloads(self, training_id: int) -> List[Dict[str, Any]]:
        existing_courses = self.db.query(TrainingCourse).filter(
            TrainingCourse.training_id == training_id
        ).order_by(TrainingCourse.id.asc()).all()
        return [self._build_course_payload(course) for course in existing_courses]

    def _persist_training_courses(self, training_id: int, courses: List[Dict[str, Any]]) -> None:
        existing_rows = self.db.query(TrainingCourse).filter(
            TrainingCourse.training_id == training_id
        ).order_by(TrainingCourse.id.asc()).all()
        row_by_key: Dict[str, TrainingCourse] = {}
        row_by_identity: Dict[str, TrainingCourse] = {}
        payload_by_row_id: Dict[int, Dict[str, Any]] = {}

        for row in existing_rows:
            existing_payload = self._build_course_payload(row)
            payload_by_row_id[row.id] = existing_payload
            course_key = str(row.course_key or existing_payload.get("course_key") or "").strip()
            if course_key:
                row_by_key[course_key] = row
            row_by_identity[self._course_identity(existing_payload)] = row

        for course in courses:
            raw_course_key = str(course.get("course_key") or "").strip()
            existing_row = row_by_key.get(raw_course_key) if raw_course_key else None
            if existing_row is None:
                existing_row = row_by_identity.get(self._course_identity(course))

            existing_payload = payload_by_row_id.get(existing_row.id) if existing_row else None
            payload = self._merge_course_payload(course, existing_payload)
            if existing_row is None:
                existing_row = TrainingCourse(training_id=training_id)
                self.db.add(existing_row)

            self._apply_course_payload(existing_row, payload)
            if existing_row.id is not None:
                payload_by_row_id[existing_row.id] = payload
            course_key = str(payload.get("course_key") or "").strip()
            if course_key:
                row_by_key[course_key] = existing_row
            row_by_identity[self._course_identity(payload)] = existing_row

    @staticmethod
    def _apply_course_payload(course: TrainingCourse, payload: Dict[str, Any]) -> None:
        course.course_id = payload.get("course_id")
        course.course_key = payload.get("course_key")
        course.name = payload["name"]
        course.location = payload.get("location")
        course.instructor = payload.get("instructor")
        course.primary_instructor_id = payload.get("primary_instructor_id")
        course.assistant_instructor_ids = payload.get("assistant_instructor_ids") or []
        course.hours = payload.get("hours") or 0
        course.type = payload.get("type") or "theory"
        course.schedules = payload.get("schedules") or []

    def _match_training_course(
        self,
        existing_courses: List[Dict[str, Any]],
        course_name: str,
        instructor_name: Optional[str],
        course_type: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        normalized_name = self._normalize_identity_text(course_name)
        candidates = [
            course for course in existing_courses
            if self._normalize_identity_text(course.get("name")) == normalized_name
        ]
        if not candidates:
            return None

        if instructor_name:
            normalized_instructor = self._normalize_identity_text(instructor_name)
            instructor_candidates = [
                course for course in candidates
                if self._normalize_identity_text(course.get("instructor")) == normalized_instructor
            ]
            if not instructor_candidates:
                raise ValueError("未找到匹配课程，请检查课程名称与教官是否一致")
            candidates = instructor_candidates

        if course_type:
            type_candidates = [course for course in candidates if (course.get("type") or "theory") == course_type]
            if not type_candidates:
                raise ValueError("未找到匹配课程，请检查课程名称与课程类型是否一致")
            candidates = type_candidates

        if len(candidates) == 1:
            return candidates[0]
        if len(candidates) > 1:
            raise ValueError("匹配到多门同名课程，请补充教官或课程类型")
        return None

    @staticmethod
    def _normalize_identity_text(value: Any) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _is_course_match_failure_reason(reason: str) -> bool:
        text = str(reason or "").strip()
        return (
            "未找到匹配课程" in text
            or "课程名称与教官是否一致" in text
            or "课程名称与课程类型是否一致" in text
            or "匹配到多门同名课程" in text
        )

    def _build_course_match_failure(
        self,
        row_number: int,
        session_seed: Dict[str, Any],
        reason: str,
    ) -> Dict[str, Any]:
        return {
            "row": row_number,
            "course_name": session_seed.get("course_name") or "",
            "instructor_name": session_seed.get("instructor_name") or "",
            "course_type": self._format_course_type_label(session_seed.get("course_type")),
            "reason": reason,
        }

    def _summarize_course_match_failures(self, failures: List[Dict[str, Any]]) -> Optional[str]:
        if not failures:
            return None
        preview = []
        for item in failures[:3]:
            course_name = item.get("course_name") or "未填写课程名称"
            instructor_name = item.get("instructor_name") or "未填写教官"
            course_type = item.get("course_type") or "未填写课程类型"
            preview.append(
                f"第 {item.get('row') or '-'} 行：课程={course_name}，教官={instructor_name}，类型={course_type}"
            )
        remaining = len(failures) - len(preview)
        suffix = f"；其余 {remaining} 行请继续核对课程、教官和课程类型" if remaining > 0 else ""
        return f"有 {len(failures)} 行课次未匹配到课程并已跳过：{'；'.join(preview)}{suffix}"

    @staticmethod
    def _schedule_exists(existing: List[Dict[str, Any]], target: Dict[str, Any]) -> bool:
        for item in existing:
            if (
                item.get("date") == target.get("date")
                and item.get("time_range") == target.get("time_range")
                and (item.get("location") or "") == (target.get("location") or "")
            ):
                return True
        return False

    def _build_course_payload(self, course: TrainingCourse) -> Dict[str, Any]:
        return {
            "course_id": course.course_id,
            "course_key": course.course_key or str(uuid.uuid4()),
            "name": course.name,
            "location": self._to_text(course.location),
            "instructor": course.instructor,
            "primary_instructor_id": course.primary_instructor_id,
            "assistant_instructor_ids": list(course.assistant_instructor_ids or []),
            "type": course.type or "theory",
            "hours": float(course.hours or 0),
            "schedules": [self._normalize_schedule_payload(item) for item in (course.schedules or [])],
        }

    def _merge_course_payload(
        self,
        course: Dict[str, Any],
        existing_course: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        existing_course = existing_course or {}
        existing_schedule_map = {
            self._schedule_identity(item): self._normalize_schedule_payload(item)
            for item in (existing_course.get("schedules") or [])
        }
        schedules = [
            self._merge_schedule_payload(schedule, existing_schedule_map.get(self._schedule_identity(schedule)))
            for schedule in (course.get("schedules") or [])
        ]
        schedules.sort(key=lambda item: (item.get("date") or "", item.get("time_range") or "", item.get("location") or ""))
        course_hours = self._normalize_course_hours(course.get("hours"))
        existing_hours = self._normalize_course_hours(existing_course.get("hours"))
        return {
            "course_key": course.get("course_key") or existing_course.get("course_key") or str(uuid.uuid4()),
            "name": course["name"],
            "location": course.get("location") or existing_course.get("location"),
            "instructor": course.get("instructor"),
            "primary_instructor_id": (
                course.get("primary_instructor_id")
                or existing_course.get("primary_instructor_id")
                or self._guess_instructor_id(course.get("instructor"))
            ),
            "assistant_instructor_ids": existing_course.get("assistant_instructor_ids") or [],
            "type": course.get("type") or "theory",
            "hours": course_hours if course_hours is not None else (existing_hours or 0),
            "schedules": schedules,
        }

    def _merge_schedule_payload(
        self,
        schedule: Dict[str, Any],
        existing_schedule: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        normalized = self._normalize_schedule_payload(schedule)
        if not existing_schedule:
            return normalized
        merged = dict(existing_schedule)
        merged.update({
            "date": normalized.get("date"),
            "time_range": normalized.get("time_range"),
            "hours": normalized.get("hours"),
            "location": normalized.get("location"),
        })
        merged["session_id"] = existing_schedule.get("session_id") or normalized.get("session_id") or str(uuid.uuid4())
        return merged

    @staticmethod
    def _normalize_course_hours(value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            return round(max(float(value), 0), 2)
        except Exception:
            return None

    def _normalize_schedule_payload(self, schedule: Dict[str, Any]) -> Dict[str, Any]:
        raw = dict(schedule or {})
        return {
            "session_id": raw.get("session_id") or raw.get("sessionId") or str(uuid.uuid4()),
            "date": raw.get("date"),
            "time_range": raw.get("time_range") or raw.get("timeRange") or "",
            "hours": round(float(raw.get("hours", 0) or 0), 2),
            "location": self._to_text(raw.get("location")),
            "status": self._to_text(raw.get("status")) or "pending",
            "started_at": raw.get("started_at") or raw.get("startedAt"),
            "checkin_started_at": raw.get("checkin_started_at") or raw.get("checkinStartedAt"),
            "checkin_ended_at": raw.get("checkin_ended_at") or raw.get("checkinEndedAt"),
            "checkout_started_at": raw.get("checkout_started_at") or raw.get("checkoutStartedAt"),
            "checkout_ended_at": raw.get("checkout_ended_at") or raw.get("checkoutEndedAt"),
            "ended_at": raw.get("ended_at") or raw.get("endedAt"),
            "skipped_at": raw.get("skipped_at") or raw.get("skippedAt"),
            "skipped_by": raw.get("skipped_by") or raw.get("skippedBy"),
            "skip_reason": raw.get("skip_reason") or raw.get("skipReason"),
        }

    @staticmethod
    def _course_identity(course: Dict[str, Any]) -> str:
        return (
            f"{str(course.get('name') or '').strip().lower()}|"
            f"{str(course.get('instructor') or '').strip().lower()}|"
            f"{str(course.get('type') or 'theory').strip().lower()}"
        )

    @staticmethod
    def _schedule_identity(schedule: Dict[str, Any]) -> str:
        return (
            f"{schedule.get('date') or ''}|"
            f"{schedule.get('time_range') or schedule.get('timeRange') or ''}|"
            f"{schedule.get('location') or ''}"
        )

    def _guess_instructor_id(self, instructor_name: Any) -> Optional[int]:
        name = self._to_text(instructor_name)
        if not name:
            return None
        user = self.db.query(User).filter(
            (User.nickname == name) | (User.username == name)
        ).first()
        return user.id if user else None

    @staticmethod
    def _normalize_course_type(value: Any) -> str:
        text = str(value or "").strip().lower()
        practice_values = {"实践", "实操", "技能", "practice", "skill"}
        return "practice" if text in practice_values else "theory"

    def _parse_optional_course_type(self, value: Any) -> Optional[str]:
        text = str(value or "").strip()
        if not text:
            return None
        return self._normalize_course_type(text)

    @staticmethod
    def _format_course_type_label(value: Optional[str]) -> str:
        if value == "practice":
            return "技能课"
        if value == "theory":
            return "理论课"
        return ""

    @staticmethod
    def _parse_date(value: Any) -> Optional[str]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (int, float)):
            try:
                base = date(1899, 12, 30)
                parsed = base + timedelta(days=int(value))
                return parsed.isoformat()
            except Exception:
                return None

        text = str(value).strip()
        if not text:
            return None
        text = text.replace("年", "-").replace("月", "-").replace("日", "")
        text = text.replace(".", "-").replace("/", "-")
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m-%d-%Y"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.date().isoformat()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_time(value: Any) -> Optional[str]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, dt_time):
            return value.strftime("%H:%M")
        if isinstance(value, (int, float)) and 0 <= float(value) < 1:
            total_minutes = int(round(float(value) * 24 * 60))
            hour = total_minutes // 60
            minute = total_minutes % 60
            return f"{hour:02d}:{minute:02d}"

        text = str(value).strip()
        if not text:
            return None
        text = text.replace("：", ":").replace(".", ":")
        text = re.sub(r"[^0-9:]", "", text)
        if re.fullmatch(r"\d{1,2}:\d{1,2}", text):
            hour, minute = text.split(":")
            h = int(hour)
            m = int(minute)
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        if re.fullmatch(r"\d{3,4}", text):
            if len(text) == 3:
                h = int(text[0])
                m = int(text[1:])
            else:
                h = int(text[:2])
                m = int(text[2:])
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        return None

    def _parse_time_range(self, value: Any) -> Tuple[Optional[str], Optional[str]]:
        text = str(value or "").strip()
        if not text:
            return None, None
        normalized = text.replace("：", ":").replace("—", "-")
        normalized = normalized.replace("至", "-").replace("到", "-").replace("~", "-")
        parts = [part.strip() for part in normalized.split("-") if part.strip()]
        if len(parts) < 2:
            return None, None
        start = self._parse_time(parts[0])
        end = self._parse_time(parts[1])
        return start, end

    def _calc_hours(self, start: str, end: str) -> float:
        start_minutes = self._minutes_from_time(start)
        end_minutes = self._minutes_from_time(end)
        if start_minutes is None or end_minutes is None:
            return 0
        delta = end_minutes - start_minutes
        if delta <= 0:
            return 0
        return round(delta / 60, 2)

    @staticmethod
    def _minutes_from_time(value: str) -> Optional[int]:
        if not value:
            return None
        try:
            hour, minute = value.split(":")
            return int(hour) * 60 + int(minute)
        except Exception:
            return None

    @staticmethod
    def _parse_float(value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except Exception:
            return None
