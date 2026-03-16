"""
系统管理 Excel 导入导出服务。
"""

from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models import Department, Permission, Role, User
from app.schemas.role import (
    ROLE_DATA_SCOPE_ALL,
    ROLE_DATA_SCOPE_CHOICES,
    ROLE_DATA_SCOPE_DEPARTMENT,
    ROLE_DATA_SCOPE_DEPARTMENT_AND_SUB,
    ROLE_DATA_SCOPE_POLICE_TYPE,
    ROLE_DATA_SCOPE_SELF,
)
from app.services.auth import auth_service
from app.services.batch_import import BatchImportService
from config import settings


USER_TEMPLATE_HEADERS = [
    "用户名",
    "密码",
    "姓名",
    "角色",
    "性别",
    "手机号",
    "警号",
    "邮箱",
]

DEPARTMENT_TEMPLATE_HEADERS = [
    "部门编码",
    "部门名称",
    "父级部门编码",
    "权限继承",
    "状态",
    "描述",
]

ROLE_TEMPLATE_HEADERS = [
    "角色编码",
    "角色名称",
    "数据范围",
    "状态",
    "描述",
]

ROLE_DATA_SCOPE_LABELS = {
    ROLE_DATA_SCOPE_ALL: "全部",
    ROLE_DATA_SCOPE_DEPARTMENT: "本部门",
    ROLE_DATA_SCOPE_DEPARTMENT_AND_SUB: "本部门及下属部门",
    ROLE_DATA_SCOPE_POLICE_TYPE: "本警种",
    ROLE_DATA_SCOPE_SELF: "本人",
}

ROLE_DATA_SCOPE_ALIASES = {
    "all": ROLE_DATA_SCOPE_ALL,
    "全部": ROLE_DATA_SCOPE_ALL,
    "department": ROLE_DATA_SCOPE_DEPARTMENT,
    "本部门": ROLE_DATA_SCOPE_DEPARTMENT,
    "departmentandsub": ROLE_DATA_SCOPE_DEPARTMENT_AND_SUB,
    "department_and_sub": ROLE_DATA_SCOPE_DEPARTMENT_AND_SUB,
    "本部门及下属部门": ROLE_DATA_SCOPE_DEPARTMENT_AND_SUB,
    "本部门及子部门": ROLE_DATA_SCOPE_DEPARTMENT_AND_SUB,
    "policetype": ROLE_DATA_SCOPE_POLICE_TYPE,
    "police_type": ROLE_DATA_SCOPE_POLICE_TYPE,
    "本警种": ROLE_DATA_SCOPE_POLICE_TYPE,
    "self": ROLE_DATA_SCOPE_SELF,
    "本人": ROLE_DATA_SCOPE_SELF,
}

ROLE_LOOKUP_ALIASES = {
    "admin": "admin",
    "管理员": "admin",
    "系统管理员": "admin",
    "instructor": "instructor",
    "教官": "instructor",
    "讲师": "instructor",
    "教师": "instructor",
    "student": "student",
    "学员": "student",
    "民警": "student",
}

BOOLEAN_TRUE_ALIASES = {
    "1",
    "true",
    "yes",
    "y",
    "是",
    "启用",
    "正常",
    "有效",
    "继承",
}

BOOLEAN_FALSE_ALIASES = {
    "0",
    "false",
    "no",
    "n",
    "否",
    "停用",
    "禁用",
    "不继承",
}


class SystemExchangeService(BatchImportService):
    """系统管理导入导出服务。"""

    USER_FIELD_ALIASES: Dict[str, set[str]] = {
        "username": {"用户名", "账号", "登录名", "user_name"},
        "password": {"密码", "登录密码", "password"},
        "name": {"姓名", "名字", "昵称", "name", "nickname"},
        "role": {"角色", "账号角色", "用户角色", "role"},
        "gender": {"性别", "gender", "sex"},
        "phone": {"手机号", "手机", "联系电话", "phone", "mobile", "tel"},
        "police_id": {"警号", "警员编号", "编号", "policeid", "badge", "badgeid"},
        "email": {"邮箱", "电子邮箱", "mail", "email"},
    }

    DEPARTMENT_FIELD_ALIASES: Dict[str, set[str]] = {
        "code": {"部门编码", "编码", "code", "departmentcode"},
        "name": {"部门名称", "名称", "name", "departmentname"},
        "parent_code": {"父级部门编码", "父部门编码", "上级部门编码", "parentcode"},
        "inherit_sub_permissions": {"权限继承", "是否继承", "inherit", "inheritsubpermissions"},
        "is_active": {"状态", "启用状态", "status", "isactive"},
        "description": {"描述", "部门描述", "remark", "description"},
    }

    ROLE_FIELD_ALIASES: Dict[str, set[str]] = {
        "code": {"角色编码", "编码", "rolecode", "code"},
        "name": {"角色名称", "名称", "rolename", "name"},
        "data_scopes": {"数据范围", "数据权限", "datascopes", "datascope"},
        "is_active": {"状态", "启用状态", "status", "isactive"},
        "description": {"描述", "角色描述", "remark", "description"},
    }

    def __init__(self, db: Session):
        super().__init__(db)
        self.header_fill = PatternFill("solid", fgColor="EAF1FF")

    # ===== 模板 =====

    def build_user_template(self) -> bytes:
        instructions = [
            ("字段", "说明"),
            ("用户名", "必填，建议与实际登录账号一致，导入更新时默认按用户名匹配"),
            ("密码", "新用户必填；已存在用户留空则不修改密码"),
            ("姓名", "选填"),
            ("角色", "选填，可填角色名称或编码；多个角色可用中文顿号、逗号分隔"),
            ("性别", "选填，可填 男/女"),
            ("手机号", "选填，可作为辅助匹配字段"),
            ("警号", "选填，可作为辅助匹配字段"),
            ("邮箱", "选填"),
        ]
        return self._build_template_bytes("用户导入", USER_TEMPLATE_HEADERS, instructions)

    def build_department_template(self) -> bytes:
        instructions = [
            ("字段", "说明"),
            ("部门编码", "必填，导入更新时按部门编码匹配"),
            ("部门名称", "必填"),
            ("父级部门编码", "选填，根部门留空；请确保父部门已存在或排在当前行之前"),
            ("权限继承", "选填，可填 是/否、继承/不继承"),
            ("状态", "选填，可填 启用/停用"),
            ("描述", "选填"),
        ]
        return self._build_template_bytes("部门导入", DEPARTMENT_TEMPLATE_HEADERS, instructions)

    def build_role_template(self) -> bytes:
        instructions = [
            ("字段", "说明"),
            ("角色编码", "必填，导入更新时按角色编码匹配；admin 不支持导入修改"),
            ("角色名称", "必填"),
            ("数据范围", "选填，多个值用逗号分隔，可填 全部/all、本部门/department、本部门及下属部门/department_and_sub、本警种/police_type、本人/self"),
            ("状态", "选填，可填 启用/停用"),
            ("描述", "选填"),
        ]
        return self._build_template_bytes("角色导入", ROLE_TEMPLATE_HEADERS, instructions)

    # ===== 导出 =====

    def export_users(self, search: Optional[str] = None, role_code: Optional[str] = None) -> bytes:
        query = self.db.query(User).options(joinedload(User.roles)).filter(User.is_active.is_(True))
        if role_code:
            query = query.filter(User.roles.any(Role.code == role_code))
        if search:
            keyword = f"%{search}%"
            query = query.filter(
                (User.nickname.ilike(keyword))
                | (User.username.ilike(keyword))
                | (User.police_id.ilike(keyword))
            )

        users = query.order_by(User.id.asc()).all()
        rows = [
            [
                user.username or "",
                "",
                user.nickname or "",
                "、".join(role.name for role in user.roles or []),
                user.gender or "",
                user.phone or "",
                user.police_id or "",
                user.email or "",
            ]
            for user in users
        ]
        return self._build_data_bytes("用户导出", USER_TEMPLATE_HEADERS, rows)

    def export_departments(
        self,
        keyword: Optional[str] = None,
        status_value: Optional[bool] = None,
        parent_value: Optional[str] = None,
    ) -> bytes:
        query = self.db.query(Department).order_by(Department.id.asc())
        if status_value is not None:
            query = query.filter(Department.is_active.is_(status_value))
        if parent_value is not None and parent_value != "":
            if parent_value == "root":
                query = query.filter(Department.parent_id.is_(None))
            else:
                try:
                    query = query.filter(Department.parent_id == int(parent_value))
                except ValueError:
                    pass

        departments = query.all()
        if keyword:
            lowered = keyword.strip().lower()
            departments = [
                item
                for item in departments
                if lowered in str(item.name or "").lower()
                or lowered in str(item.code or "").lower()
                or lowered in str(item.description or "").lower()
            ]

        all_department_code_map = {
            item.id: item.code for item in self.db.query(Department.id, Department.code).all()
        }
        rows = []
        for department in self._order_departments(departments):
            parent_code = all_department_code_map.get(department.parent_id) or ""
            rows.append(
                [
                    department.code or "",
                    department.name or "",
                    parent_code,
                    "是" if department.inherit_sub_permissions else "否",
                    "启用" if department.is_active else "停用",
                    department.description or "",
                ]
            )
        return self._build_data_bytes("部门导出", DEPARTMENT_TEMPLATE_HEADERS, rows)

    def export_roles(
        self,
        name: Optional[str] = None,
        is_active: Optional[bool] = None,
    ) -> bytes:
        query = self.db.query(Role).order_by(Role.created_at.asc(), Role.id.asc())
        if name:
            query = query.filter(Role.name.contains(name))
        if is_active is not None:
            query = query.filter(Role.is_active.is_(is_active))

        rows = []
        for role in query.all():
            rows.append(
                [
                    role.code or "",
                    role.name or "",
                    "、".join(ROLE_DATA_SCOPE_LABELS.get(item, item) for item in role.data_scopes or []),
                    "启用" if role.is_active else "停用",
                    role.description or "",
                ]
            )
        return self._build_data_bytes("角色导出", ROLE_TEMPLATE_HEADERS, rows)

    # ===== 导入 =====

    def import_users(self, file_bytes: bytes, default_role_code: str = "student") -> Dict[str, Any]:
        rows = self._read_excel_rows(file_bytes, self.USER_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        summary = {
            "total_rows": 0,
            "success_rows": 0,
            "created_count": 0,
            "updated_count": 0,
            "password_reset_count": 0,
            "skipped_count": 0,
            "skipped_rows": [],
        }

        for row_number, row in rows:
            summary["total_rows"] += 1
            try:
                with self.db.begin_nested():
                    result = self._upsert_user_from_system_row(row, default_role_code)
                    if result["created"]:
                        summary["created_count"] += 1
                    if result["updated"]:
                        summary["updated_count"] += 1
                    if result["password_reset"]:
                        summary["password_reset_count"] += 1
                    summary["success_rows"] += 1
            except Exception as exc:
                summary["skipped_rows"].append({"row": row_number, "reason": str(exc)})

        self.db.commit()
        summary["skipped_count"] = len(summary["skipped_rows"])
        return summary

    def import_departments(self, file_bytes: bytes) -> Dict[str, Any]:
        rows = self._read_excel_rows(file_bytes, self.DEPARTMENT_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        summary = {
            "total_rows": 0,
            "success_rows": 0,
            "created_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "skipped_rows": [],
        }

        for row_number, row in rows:
            summary["total_rows"] += 1
            try:
                with self.db.begin_nested():
                    result = self._upsert_department_from_row(row)
                    if result["created"]:
                        summary["created_count"] += 1
                    if result["updated"]:
                        summary["updated_count"] += 1
                    summary["success_rows"] += 1
            except Exception as exc:
                summary["skipped_rows"].append({"row": row_number, "reason": str(exc)})

        self.db.commit()
        summary["skipped_count"] = len(summary["skipped_rows"])
        return summary

    def import_roles(self, file_bytes: bytes) -> Dict[str, Any]:
        rows = self._read_excel_rows(file_bytes, self.ROLE_FIELD_ALIASES)
        if not rows:
            raise ValueError("Excel文件为空，或缺少可识别的表头")

        summary = {
            "total_rows": 0,
            "success_rows": 0,
            "created_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "skipped_rows": [],
        }

        for row_number, row in rows:
            summary["total_rows"] += 1
            try:
                with self.db.begin_nested():
                    result = self._upsert_role_from_row(row)
                    if result["created"]:
                        summary["created_count"] += 1
                    if result["updated"]:
                        summary["updated_count"] += 1
                    summary["success_rows"] += 1
            except Exception as exc:
                summary["skipped_rows"].append({"row": row_number, "reason": str(exc)})

        self.db.commit()
        summary["skipped_count"] = len(summary["skipped_rows"])
        return summary

    # ===== 行处理 =====

    def _upsert_user_from_system_row(
        self,
        row: Dict[str, Any],
        default_role_code: str,
    ) -> Dict[str, bool]:
        username = self._to_text(row.get("username"))
        password = self._to_text(row.get("password"))
        nickname = self._to_text(row.get("name"))
        gender = self._normalize_gender(row.get("gender"))
        phone = self._normalize_phone(row.get("phone"))
        police_id = self._to_text(row.get("police_id"))
        email = self._to_text(row.get("email"))
        role_values = self._split_multi_values(row.get("role"))

        if not username:
            raise ValueError("用户名不能为空")

        user = self._find_user(police_id=police_id, username=username, phone=phone)
        if user:
            if username != user.username:
                raise ValueError("匹配到已有用户后不允许通过导入修改用户名")

            updated = False
            password_reset = False

            if password:
                user.password_hash = auth_service.get_password_hash(password)
                updated = True
                password_reset = True

            if nickname is not None and nickname != (user.nickname or ""):
                user.nickname = nickname
                updated = True

            if gender is not None and gender != (user.gender or ""):
                user.gender = gender
                updated = True

            if phone and phone != (user.phone or ""):
                self._ensure_unique_user_field("phone", phone, user.id, "手机号")
                user.phone = phone
                updated = True

            if police_id and police_id != (user.police_id or ""):
                self._ensure_unique_user_field("police_id", police_id, user.id, "警号")
                user.police_id = police_id
                updated = True

            if email and email != (user.email or ""):
                self._ensure_unique_user_field("email", email, user.id, "邮箱")
                user.email = email
                updated = True

            if role_values and not self._is_protected_admin_user(user):
                roles = self._resolve_roles_by_text(role_values)
                if self._replace_user_roles(user, roles):
                    updated = True

            self.db.flush()
            return {"created": False, "updated": updated, "password_reset": password_reset}

        if not password:
            raise ValueError("新用户必须填写密码")

        roles = self._resolve_roles_by_text(role_values or [default_role_code])
        self._ensure_unique_user_field("username", username, None, "用户名")
        if phone:
            self._ensure_unique_user_field("phone", phone, None, "手机号")
        if police_id:
            self._ensure_unique_user_field("police_id", police_id, None, "警号")
        if email:
            self._ensure_unique_user_field("email", email, None, "邮箱")

        new_user = User(
            username=username,
            password_hash=auth_service.get_password_hash(password),
            nickname=nickname,
            gender=gender,
            phone=phone,
            police_id=police_id,
            email=email,
            is_active=True,
        )
        new_user.roles = roles
        self.db.add(new_user)
        self.db.flush()
        return {"created": True, "updated": True, "password_reset": False}

    def _upsert_department_from_row(self, row: Dict[str, Any]) -> Dict[str, bool]:
        code = self._to_text(row.get("code"))
        name = self._to_text(row.get("name"))
        parent_code = self._to_text(row.get("parent_code"))
        inherit_sub_permissions = self._parse_bool_value(
            row.get("inherit_sub_permissions"),
            field_label="权限继承",
        )
        is_active = self._parse_bool_value(row.get("is_active"), field_label="状态")
        description = self._to_text(row.get("description"))

        if not code:
            raise ValueError("部门编码不能为空")
        if not name:
            raise ValueError("部门名称不能为空")

        department = self.db.query(Department).filter(Department.code == code).first()
        created = False
        updated = False

        if department is None:
            department = Department(
                code=code,
                name=name,
                inherit_sub_permissions=inherit_sub_permissions if inherit_sub_permissions is not None else False,
                is_active=is_active if is_active is not None else True,
                description=description,
            )
            self.db.add(department)
            self.db.flush()
            created = True
            updated = True
        else:
            if name != department.name:
                department.name = name
                updated = True
            if inherit_sub_permissions is not None and inherit_sub_permissions != department.inherit_sub_permissions:
                department.inherit_sub_permissions = inherit_sub_permissions
                updated = True
            if is_active is not None and is_active != department.is_active:
                department.is_active = is_active
                updated = True
            if description != (department.description or ""):
                department.description = description
                updated = True

        parent_id = None
        if parent_code:
            parent_department = self.db.query(Department).filter(Department.code == parent_code).first()
            if parent_department is None:
                raise ValueError(f"父级部门不存在: {parent_code}")
            if parent_department.id == department.id:
                raise ValueError("部门不能设置自己为父级部门")
            if department.id in self._get_ancestor_department_ids(parent_department.id):
                raise ValueError("设置父级部门会形成循环引用")
            parent_id = parent_department.id

        if department.parent_id != parent_id:
            department.parent_id = parent_id
            updated = True

        self.db.flush()
        return {"created": created, "updated": updated}

    def _upsert_role_from_row(self, row: Dict[str, Any]) -> Dict[str, bool]:
        code = self._to_text(row.get("code"))
        name = self._to_text(row.get("name"))
        data_scopes = self._parse_role_data_scopes(row.get("data_scopes"))
        is_active = self._parse_bool_value(row.get("is_active"), field_label="状态")
        description = self._to_text(row.get("description"))

        if not code:
            raise ValueError("角色编码不能为空")
        if not name:
            raise ValueError("角色名称不能为空")

        role = self.db.query(Role).options(joinedload(Role.permissions)).filter(Role.code == code).first()
        created = False
        updated = False

        if role is None:
            if code == "admin":
                raise ValueError("admin 角色不支持通过导入创建")

            role = Role(
                code=code,
                name=name,
                description=description,
                is_active=is_active if is_active is not None else True,
                data_scopes=self._normalize_role_data_scopes(code, data_scopes),
            )
            role.permissions = self._get_fixed_permissions()
            self.db.add(role)
            self.db.flush()
            created = True
            updated = True
            return {"created": created, "updated": updated}

        if role.code == "admin":
            raise ValueError("admin 角色不支持通过导入修改")

        if name != role.name:
            role.name = name
            updated = True

        normalized_scopes = self._normalize_role_data_scopes(role.code, data_scopes)
        if normalized_scopes != list(role.data_scopes or []):
            role.data_scopes = normalized_scopes
            updated = True

        if is_active is not None and is_active != role.is_active:
            role.is_active = is_active
            updated = True

        if description != (role.description or ""):
            role.description = description
            updated = True

        for permission in self._get_fixed_permissions():
            if permission.id not in {item.id for item in role.permissions or []}:
                role.permissions.append(permission)
                updated = True

        self.db.flush()
        return {"created": created, "updated": updated}

    # ===== Workbook =====

    def _build_template_bytes(
        self,
        title: str,
        headers: Sequence[str],
        instructions: Sequence[Sequence[str]],
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title
        self._append_headers(sheet, headers)

        help_sheet = workbook.create_sheet("填写说明")
        for row in instructions:
            help_sheet.append(list(row))
        self._style_sheet(help_sheet)
        self._autosize_columns(help_sheet)
        return self._workbook_to_bytes(workbook)

    def _build_data_bytes(
        self,
        title: str,
        headers: Sequence[str],
        rows: Iterable[Sequence[Any]],
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title
        self._append_headers(sheet, headers)
        for row in rows:
            sheet.append(list(row))
        self._style_sheet(sheet)
        self._autosize_columns(sheet)
        return self._workbook_to_bytes(workbook)

    def _append_headers(self, sheet, headers: Sequence[str]) -> None:
        sheet.append(list(headers))
        self._style_sheet(sheet)

    def _style_sheet(self, sheet) -> None:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    def _autosize_columns(self, sheet) -> None:
        widths: Dict[int, int] = {}
        for row in sheet.iter_rows(values_only=True):
            for index, value in enumerate(row, start=1):
                text = str(value or "")
                widths[index] = max(widths.get(index, 12), min(max(len(text) + 2, 12), 48))

        for index, width in widths.items():
            sheet.column_dimensions[get_column_letter(index)].width = width

    @staticmethod
    def _workbook_to_bytes(workbook: Workbook) -> bytes:
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    # ===== Helper =====

    def _resolve_roles_by_text(self, values: List[str]) -> List[Role]:
        lookup: Dict[str, Role] = {}
        for role in self.db.query(Role).all():
            lookup[self._normalize_header(role.code)] = role
            lookup[self._normalize_header(role.name)] = role

        result: List[Role] = []
        missing: List[str] = []
        seen_ids: set[int] = set()

        for raw_value in values:
            normalized = self._normalize_header(raw_value)
            normalized = ROLE_LOOKUP_ALIASES.get(normalized, normalized)
            role = lookup.get(normalized)
            if role is None:
                missing.append(str(raw_value))
                continue
            if role.id not in seen_ids:
                seen_ids.add(role.id)
                result.append(role)

        if missing:
            raise ValueError(f"角色不存在: {','.join(missing)}")
        return result

    def _parse_role_data_scopes(self, value: Any) -> List[str]:
        scopes: List[str] = []
        for raw_item in self._split_multi_values(value):
            normalized = ROLE_DATA_SCOPE_ALIASES.get(self._normalize_header(raw_item))
            if normalized is None:
                raise ValueError(f"不支持的数据范围: {raw_item}")
            if normalized not in ROLE_DATA_SCOPE_CHOICES:
                raise ValueError(f"不支持的数据范围: {raw_item}")
            if normalized not in scopes:
                scopes.append(normalized)
        return scopes

    def _normalize_role_data_scopes(self, role_code: str, data_scopes: List[str]) -> List[str]:
        if role_code == "admin":
            return [ROLE_DATA_SCOPE_ALL]
        normalized: List[str] = []
        for item in data_scopes:
            if item not in normalized:
                normalized.append(item)
        return normalized

    def _parse_bool_value(self, value: Any, field_label: str) -> Optional[bool]:
        text = self._to_text(value)
        if text is None:
            return None
        normalized = self._normalize_header(text)
        if normalized in BOOLEAN_TRUE_ALIASES:
            return True
        if normalized in BOOLEAN_FALSE_ALIASES:
            return False
        raise ValueError(f"{field_label}不支持的值: {text}")

    def _ensure_unique_user_field(
        self,
        field_name: str,
        value: Optional[str],
        exclude_user_id: Optional[int],
        label: str,
    ) -> None:
        if not value:
            return
        field = getattr(User, field_name)
        query = self.db.query(User.id).filter(field == value)
        if exclude_user_id is not None:
            query = query.filter(User.id != exclude_user_id)
        if query.first():
            raise ValueError(f"{label}已存在: {value}")

    def _replace_user_roles(self, user: User, roles: List[Role]) -> bool:
        current_ids = {role.id for role in user.roles or []}
        target_ids = {role.id for role in roles}
        if current_ids == target_ids:
            return False
        user.roles = roles
        return True

    @staticmethod
    def _is_protected_admin_user(user: User) -> bool:
        return str(user.username or "").lower() == "admin"

    def _get_fixed_permissions(self) -> List[Permission]:
        fixed_permissions = self.db.query(Permission).filter(
            Permission.code.in_(settings.FIXED_PERMISSIONS)
        ).all()
        deduplicated: Dict[int, Permission] = {}
        for permission in fixed_permissions:
            deduplicated[permission.id] = permission
        return list(deduplicated.values())

    def _get_ancestor_department_ids(self, department_id: int) -> List[int]:
        department_ids: List[int] = []
        current_department_id: Optional[int] = department_id

        while current_department_id is not None:
            department_ids.append(current_department_id)
            current_department = self.db.query(Department).filter(
                Department.id == current_department_id
            ).first()
            if not current_department or current_department.parent_id is None:
                break
            current_department_id = current_department.parent_id

        return department_ids

    def _order_departments(self, departments: List[Department]) -> List[Department]:
        node_map = {item.id: item for item in departments}
        children_map: Dict[Optional[int], List[Department]] = {}
        for item in departments:
            children_map.setdefault(item.parent_id, []).append(item)

        for child_list in children_map.values():
            child_list.sort(key=lambda item: (item.name or "", item.id))

        ordered: List[Department] = []

        def walk(parent_id: Optional[int]) -> None:
            for child in children_map.get(parent_id, []):
                ordered.append(child)
                walk(child.id)

        walk(None)
        remaining_ids = {item.id for item in departments} - {item.id for item in ordered}
        for item in departments:
            if item.id in remaining_ids:
                ordered.append(item)
        return ordered
